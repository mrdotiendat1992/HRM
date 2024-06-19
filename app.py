from flask import Flask, render_template, request, url_for, redirect, flash, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, current_user, login_required
from flask_paginate import Pagination, get_page_parameter
import pyodbc
import openpyxl
import pandas as pd
from datetime import datetime, timedelta
from database import *
import os
from werkzeug.utils import secure_filename
import re
from functools import wraps

app = Flask(__name__)
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///db.sqlite"
app.config["SECRET_KEY"] = "abc"
app.config['UPLOAD_FOLDER'] = r'./static/uploads'
db = SQLAlchemy()
 
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
 
used_db = "Driver={SQL Server};Server=172.16.60.100;Database=HR;UID=huynguyen;PWD=Namthuan@123;"
#used_db = "Driver={SQL Server}; Server=DESKTOP-G635SF6; Database=HR; Trusted_Connection=yes;"
print(used_db)

class Users(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    masothe = db.Column(db.String(250), nullable=False)
    hoten = db.Column(db.String(250), nullable=False)
    macongty = db.Column(db.String(250), nullable=False)
    tencongty = db.Column(db.String(250), nullable=False)
    phongban = db.Column(db.String(250), nullable=False)
    matkhau = db.Column(db.String(250), nullable=False)      
    role = db.Column(db.String(250), nullable=False)  
 
db.init_app(app)
 
 
with app.app_context():
    db.create_all()
 
@login_manager.user_loader
def loader_user(user_id):
    return Users.query.get(int(user_id))

# Role-based decorator
def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated or current_user.role not in roles:
                return redirect(url_for('unauthorized'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def dieuchuyennhansu(mst,
                    loaidieuchuyen,
                    vitricu,
                    vitrimoi,
                    chuyencu,
                    chuyenmoi,
                    gradecodecu,
                    gradecodemoi,
                    sectioncodecu,
                    sectioncodemoi,
                    hccategorycu,
                    hccategorymoi,
                    departmentcu,
                    departmentmoi,
                    sectiondescriptioncu,
                    sectiondescriptionmoi,
                    employeetypecu,
                    employeetypemoi,
                    positioncodedescriptioncu,
                    positioncodedescriptionmoi,
                    positioncodecu,
                    positioncodemoi,
                    vitriencu,
                    vitrienmoi,
                    ngaydieuchuyen
                   ):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query1 = f"INSERT INTO HR.dbo.Lich_su_cong_tac VALUES ('{current_user.macongty}','{mst}','{chuyencu}',N'{vitricu}','{chuyenmoi}',N'{vitrimoi}',N'{loaidieuchuyen}','{ngaydieuchuyen}')"
    print(query1)
    cursor.execute(query1)
    query2 = f"UPDATE HR.dbo.Danh_sach_CBCNV SET Job_title_VN = N'{vitrimoi}', Line = '{chuyenmoi}', Headcount_category = '{hccategorymoi}', Department = '{departmentmoi}', Section_description = '{sectiondescriptionmoi}', Emp_type = '{employeetypemoi}', Position_code_description = '{positioncodedescriptionmoi}', Section_code = '{sectioncodemoi}', Grade_code = '{gradecodemoi}', Position_code = '{positioncodemoi}', Job_title_EN = N'{vitrienmoi}' WHERE MST = '{mst}' AND Factory = '{current_user.macongty}'"
    print(query2)
    cursor.execute(query2)
    conn.commit()
    conn.close()
    
def dichuyennghiviec(mst,
                    loaidieuchuyen,
                    vitricu,
                    vitrimoi,
                    chuyencu,
                    chuyenmoi,
                    gradecodecu,
                    gradecodemoi,
                    sectioncodecu,
                    sectioncodemoi,
                    hccategorycu,
                    hccategorymoi,
                    departmentcu,
                    departmentmoi,
                    sectiondescriptioncu,
                    sectiondescriptionmoi,
                    employeetypecu,
                    employeetypemoi,
                    positioncodedescriptioncu,
                    positioncodedescriptionmoi,
                    positioncodecu,
                    positioncodemoi,
                    vitriencu,
                    vitrienmoi,
                    ngaydieuchuyen
                   ):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query1 = f"INSERT INTO HR.dbo.Lich_su_cong_tac VALUES ('{current_user.macongty}','{mst}','{chuyencu}',N'{vitricu}',NULL,NULL,N'Nghỉ việc','{ngaydieuchuyen}')"   
    print(query1)
    cursor.execute(query1)
    query2 = f"UPDATE HR.dbo.Danh_sach_CBCNV SET Job_title_VN = NULL, Line = NULL, Headcount_category = NULL, Department = NULL, Section_description = NULL, Emp_type = NULL, Position_code_description = NULL, Section_code = NULL, Grade_code = NULL, Position_code = NULL, Job_title_EN = NULL, Trang_thai_lam_viec = N'Nghỉ việc', Ngay_nghi = '{ngaydieuchuyen}' WHERE MST = '{mst}' AND Factory = '{current_user.macongty}'"
    print(query2)
    cursor.execute(query2)
    conn.commit()
    conn.close()


def thaydoithongtinhopdong(kieuhopdong,mst,ngaylamhopdong,thanglamhopdong,namlamhopdong,ngayketthuchopdong,thangketthuchopdong,namketthuchopdong,tennhanvien,ngaysinh,gioitinh,thuongtru,cccd,ngaycapcccd,mucluong,chucvu,bophan):
    
    try:
        print(kieuhopdong)
        if kieuhopdong == "HĐ thử việc":
            if current_user.macongty == "NT1":
                try:
                    workbook = openpyxl.load_workbook(f'HĐLĐ_NT1/HĐLĐ THỬ VIỆC.xlsx')
                    sheet = workbook.active  # or workbook['SheetName']

                    # Change the value of a specific cell
                    sheet['E4'] = f'Số: PC/{mst}'
                    sheet['M4'] = f'Hải Phòng, ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}'
                    sheet['D18'] = tennhanvien
                    sheet['E19'] = ngaysinh
                    sheet['Q19'] = gioitinh
                    sheet['F20'] = thuongtru
                    sheet['B21'] = f"Số CCCD:{cccd}"
                    sheet['L21'] = ngaycapcccd
                    sheet['B25'] = f"Từ ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong} đến hết ngày {ngayketthuchopdong} tháng {thangketthuchopdong} năm {namketthuchopdong}"
                    sheet['G31'] = chucvu
                    sheet['G42'] = f"{mucluong} VNĐ/tháng"        
                    filepath = f'NT1 - Hợp đồng thử việc - {mst} - {ngaylamhopdong}{thanglamhopdong}{namlamhopdong}.xlsx'
                    workbook.save(filepath)
                    print(filepath)
                    return filepath
                except Exception as e:
                    print(e)
                    return None
            elif current_user.macongty == "NT2":
                try:
                    workbook = openpyxl.load_workbook(f'HĐLĐ_NT2/HĐLĐ THỬ VIỆC.xlsx')
                    sheet = workbook.active  # or workbook['SheetName']

                    # Change the value of a specific cell
                    sheet['E4'] = f'Số: PC/{mst}'
                    sheet['M4'] = f'Nghệ An, ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}'
                    sheet['D19'] = tennhanvien
                    sheet['E21'] = ngaysinh
                    sheet['E20'] = gioitinh
                    sheet['F22'] = thuongtru
                    sheet['D24'] = f"'{cccd}"
                    sheet['L24'] = ngaycapcccd
                    sheet['B28'] = f"Từ ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong} đến hết ngày {ngayketthuchopdong} tháng {thangketthuchopdong} năm {namketthuchopdong}"
                    sheet['G42'] = f"{mucluong} VNĐ/tháng"    
                    sheet['F33'] = chucvu
                    sheet['L33'] = chucvu   
                    sheet['F34'] = bophan
                    sheet['F35'] = f"'{chucvu}"
                    filepath = f'NT2 - Hợp đồng thử việc - {mst} - {ngaylamhopdong}{thanglamhopdong}{namlamhopdong}.xlsx'
                    workbook.save(filepath)
                    print(filepath)
                    return filepath
                except Exception as e:
                    print(e)
                    return None
        elif kieuhopdong == "HĐ có thời hạn 1 năm":
            try:
                
                workbook = openpyxl.load_workbook('HĐLĐ_NT1/HĐLĐ 1 NĂM.xlsx')
                sheet = workbook.active  # or workbook['SheetName']

                # Change the value of a specific cell
                sheet['E4'] = f'Số: LC12/{mst}'
                sheet['M4'] = f'Hải Phòng, ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}'
                sheet['D18'] = tennhanvien
                sheet['E19'] = ngaysinh
                sheet['Q19'] = gioitinh
                sheet['F20'] = thuongtru
                sheet['B21'] = f"Số CCCD:{cccd}"
                sheet['L21'] = ngaycapcccd
                sheet['B25'] = f"Từ ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong} đến hết ngày {ngayketthuchopdong} tháng {thangketthuchopdong} năm {namketthuchopdong}"
                sheet['G38'] = f"{mucluong} VNĐ/tháng"
                
                filepath = f'Hợp đồng có thời hạn 12 tháng - {mst} - {ngaylamhopdong}{thanglamhopdong}{namlamhopdong}.xlsx'
                workbook.save(filepath)
                print(filepath)
                return filepath
            except Exception as e:
                print(e)
                return None
        elif kieuhopdong == "HĐ vô thời hạn":
            try:
                workbook = openpyxl.load_workbook('HĐLĐ_NT1/HĐLĐ VÔ THỜI HẠN.xlsx')
                sheet = workbook.active  # or workbook['SheetName']

                # Change the value of a specific cell
                sheet['E4'] = f'Số: LC/{mst}'
                sheet['M4'] = f'Hải Phòng, ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}'
                sheet['D18'] = tennhanvien
                sheet['E19'] = ngaysinh
                sheet['Q19'] = gioitinh
                sheet['F20'] = thuongtru
                sheet['B21'] = f"Số CCCD:{cccd}"
                sheet['L21'] = ngaycapcccd
                sheet['B25'] = f"Kể từ ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}"
                sheet['G38'] = f"{mucluong} VNĐ/tháng"        
                filepath = f'Hợp đồng không thời hạn - {mst} - {ngaylamhopdong}{thanglamhopdong}{namlamhopdong}.xlsx'
                workbook.save(filepath)
                print(filepath)
                return filepath
            except Exception as e:
                print(e)
                return None
        elif kieuhopdong == "Chấm dứt hợp đồng":
            try:
                workbook = openpyxl.load_workbook('HĐLĐ_NT1/CHẤM DỨT HĐLĐ.xlsx')
            except Exception as e:
                print(e)
                return None        
    except Exception as e:
        return None
    
def inchamduthd(mst,
                ngaylamhopdong,
                thanglamhopdong,
                namlamhopdong,
                tennhanvien,
                chucvu,
                ngaynghi):
    
    try:
        workbook = openpyxl.load_workbook(f'HĐLĐ_NT1/CHẤM DỨT HĐ.xlsx')
        sheet = workbook.active  # or workbook['SheetName']

        # Change the value of a specific cell
        sheet['C4'] = f'{mst}'
        sheet['H5'] = f'Hải Phòng, ngày {ngaylamhopdong} tháng {thanglamhopdong} năm {namlamhopdong}'
        sheet['I12'] = f'{mst}'
        sheet['G16'] = tennhanvien
        sheet['C21'] = tennhanvien
        sheet['B26'] = tennhanvien
        sheet['D19'] = ngaynghi
        sheet['E22'] = ngaynghi
        sheet['D17'] = chucvu     
        filepath = f'Chấm dứt hợp đồng - {mst} - {ngaylamhopdong}{thanglamhopdong}{namlamhopdong}.xlsx'
        workbook.save(filepath)
        print(filepath)
        return filepath
    except Exception as e:
        print(e)
        return None
   
def laylichsucongtactheomst(mst):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query= f"SELECT * FROM HR.dbo.Lich_su_cong_tac WHERE MST = '{mst}' AND Nha_may = '{current_user.macongty}'"
    print(query)
    rows = cursor.execute(query)
    result = []
    for row in rows:
        result.append({
            "Nhà máy": row[0],
            "MST": row[1],
            "Chuyền cũ": row[2],
            "Chuyền mới": row[4],
            "Vị trí cũ": row[3],
            "Vị trí mới": row[5],
            "Phân loại": row[6],
            "Ngày thực hiện": row[7]
        })
    conn.commit()
    conn.close()
    return result
                     
def laydanhsachlinetheovitri(vitri):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Line FROM HR.dbo.HC_Name WHERE Factory = '{current_user.macongty}' AND Detail_job_title_VN = N'{vitri}'"
    print(query)
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row[0])
    return result

def lay_user(user):
    if user:
        return {
            "MST": user[0],
            "Thẻ chấm công": user[1],
            "Họ tên": user[2],
            "Số điện thoại": user[3],
            "Ngày sinh": user[4],
            "Giới tính": user[5],
            "CCCD": user[6],
            "Ngày cấp CCCD": user[7],
            "Nơi cấp": user[8],
            "CMT": user[9],
            "Thường trú": user[10],
            "Thôn xóm": user[11],
            "Phường xã": user[12],
            "Quận huyện": user[13],
            "Tỉnh thành phố": user[14],
            "Dân tộc": user[15],
            "Quốc tịch": user[16],
            "Tôn giáo": user[17],
            "Học vấn": user[18],
            "Nơi sinh": user[19],
            "Tạm trú": user[20],
            "Số BHXH": user[21],
            "Mã số thuế": user[22],
            "Ngân hàng": user[23],
            "Số tài khoản": user[24],
            "Con nhỏ": user[25],
            "Tên con 1": user[26],
            "Ngày sinh con 1": user[27],
            "Tên con 2": user[28],
            "Ngày sinh con 2": user[29],
            "Tên con 3": user[30],
            "Ngày sinh con 3": user[31],
            "Tên con 4": user[32],
            "Ngày sinh con 4": user[33],
            "Tên con 5": user[34],
            "Ngày sinh con 5": user[35],
            "Ảnh chân dung": user[36],
            "Người thân": user[39],
            "SĐT liên hệ": user[38],
            "Loại hợp đồng": user[39],
            "Ngày ký HĐ": user[40],
            "Ngày hết hạn": user[41],
            "Job title VN": user[42],
            "HC category": user[43],
            "Gradecode": user[44],
            "Factory": user[45],
            "Department": user[46],
            "Chức vụ": user[47],
            "Section code": user[48],
            "Section description": user[49],
            "Line": user[50],
            "Employee type": user[51],
            "Job title EN": user[52],
            "Position code": user[53],
            "Position description": user[54],
            "Lương cơ bản": user[55],
            "Phụ cấp": user[56],
            "Tiền phụ cấp": user[57],
            "Ngày vào": user[58],
            "Ngày nghỉ": user[59],
            "Trạng thái": user[60],
            "Ngày vào nối thâm niên": user[61],
            "Mật khẩu": user[62],
        }
    else:
        return None

def laydanhsachuser(mst, hoten, sdt, cccd, gioitinh, ngayvao, ngaynghi, ngaykyhd, ngayhethanhd, phongban, trangthai, hccategory):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE Factory = '{current_user.macongty}'"
    if mst:
        query += f" AND MST = '{mst}'"
    if hoten:
        query += f" AND Ho_ten like N'%{hoten}%'"
    if sdt:
        query += f" AND SDT = '{sdt}'"
    if cccd:
        query += f" AND CCCD = '{cccd}'"
    if gioitinh:
        query += f" AND Gioi_tinh = N'{gioitinh}'"
    if ngayvao:
        query += f" AND Ngay_vao = '{ngayvao}'"
    if ngaynghi:
        query += f" AND Ngay_nghi = '{ngaynghi}'"
    if ngaykyhd:
        query += f" AND Ngay_ky_HD = '{ngaykyhd}'"
    if ngayhethanhd:
        query += f" AND Ngay_het_han_HD = '{ngayhethanhd}'"
    if phongban:
        query += f" AND Department = '{phongban}'"
    if trangthai:
        query += f" AND Trang_thai_lam_viec = N'{trangthai}'"
    if hccategory:
        query += f" AND Headcount_category = '{hccategory}'"
    print(query)
    query += " ORDER BY CAST(mst AS INT) ASC"
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laycacphongban():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Department FROM HR.dbo.HC_Name WHERE Factory = '{current_user.macongty}'"
    print(query)
    cacphongban =  cursor.execute(query).fetchall()
    conn.close()
    result = []
    for x in cacphongban:
        result.append(x[0])
    return result

def laycacto():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Line FROM HR.dbo.HC_Name WHERE Factory = '{current_user.macongty}'"
    print(query)
    cacto =  cursor.execute(query).fetchall()
    conn.close()
    return [x[0] for x in cacto]

def laycachccategory():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT HC_category FROM HR.dbo.HC_Name WHERE Factory = '{current_user.macongty}'"
    print(query)
    cachccategory =  cursor.execute(query).fetchall()
    conn.close()
    return [x[0] for x in cachccategory]


def laydanhsachtheomst(mst):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE MST = '{mst}' AND Factory = '{current_user.macongty}'"
    print(query)
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laydanhsachusertheophongban(phongban):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE Department = '{phongban}' AND Factory = '{current_user.macongty}' ORDER BY CAST(mst AS INT) ASC"
    print(query)
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laydanhsachusertheogioitinh(gioitinh):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE Gioi_tinh = N'{gioitinh}' AND Factory = '{current_user.macongty}' ORDER BY CAST(mst AS INT) ASC"
    print(query)
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laydanhsachusertheoline(line):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE Line = '{line} 'AND Factory = '{current_user.macongty}' ORDER BY CAST(mst AS INT) ASC"
    print(query)
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laydanhsachusertheostatus(status):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Danh_sach_CBCNV WHERE Trang_thai_lam_viec = N'{status}' AND Factory = '{current_user.macongty}' ORDER BY CAST(mst AS INT) ASC"
    print(query)
    users = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for user in users:
        result.append(lay_user(user))
    return result

def laycactrangthai():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Trang_thai_lam_viec FROM HR.dbo.Danh_sach_CBCNV WHERE Factory = '{current_user.macongty}'"
    print(query)
    cactrangtha =  cursor.execute(query).fetchall()
    conn.close()
    return [x[0] for x in cactrangtha]

def laycacvitri():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Detail_job_title_VN FROM HR.dbo.HC_Name WHERE Factory = '{current_user.macongty}'"
    print(query)
    cacvitri =  cursor.execute(query).fetchall()
    conn.close()
    return [x[0] for x in cacvitri]

def laycacca():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT DISTINCT Ten_ca FROM HR.dbo.Ca_lam_viec"
    print(query)
    cacca =  cursor.execute(query).fetchall()
    conn.close()
    return [x[0] for x in cacca]

def layhcname(jobtitle,line):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query =  f"SELECT * FROM HR.dbo.HC_Name WHERE Detail_job_title_VN = N'{jobtitle}' AND Line = N'{line}' AND Factory = N'{current_user.macongty}'"
    # print(query)
    result = cursor.execute(query).fetchone()
    conn.close()
    return result

def laydanhsachdangkytuyendung(sdt=None, cccd=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if not sdt and not cccd:
        query = f"SELECT * FROM HR.dbo.Dang_ky_thong_tin WHERE Nha_may = '{current_user.macongty}'"
    elif sdt and not cccd:
        query = f"SELECT * FROM HR.dbo.Dang_ky_thong_tin WHERE Nha_may = '{current_user.macongty}' AND Sdt = '{sdt}'"
    elif not sdt and cccd:
        query = f"SELECT * FROM HR.dbo.Dang_ky_thong_tin WHERE Nha_may = '{current_user.macongty}' AND CCCD = '{cccd}'"
    else:
        query = f"SELECT * FROM HR.dbo.Dang_ky_thong_tin WHERE Nha_may = '{current_user.macongty}' AND Sdt = '{sdt}' AND CCCD = '{cccd}'"
    query+= " ORDER BY Ngay_gui_thong_tin DESC"
    print(query)
    rows =  cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append({
            "Nhà máy": row[0],
            "Vị trí tuyển dụng": row[1],
            "Họ tên": row[2],
            "Số điện thoại": row[3],
            "CCCD": row[4],
            "Dân tộc": row[5],
            "Tôn giáo": row[6],
            "Quốc tịch": row[7],
            "Học vấn": row[8],
            "Nơi sinh": row[9],
            "Tạm trú": row[10],
            "Số BHXH": row[11],
            "Mã số thuế": row[12],
            "Ngân hàng": row[13],
            "Số tài khoản": row[14],
            "Tên người thân": row[15],
            "SĐT người thân": row[16],
            "Kênh tuyển dụng": row[17],
            "Kinh nghiệm": row[18],
            "Mức lương": row[19],
            "Ngày nhận việc": row[20],
            "Có con nhỏ": row[21],
            "Tên con 1": row[22],
            "Ngày sinh con 1": row[23],
            "Tên con 2": row[24],
            "Ngày sinh con 2": row[25],
            "Tên con 3": row[26],
            "Ngày sinh con 3": row[27],
            "Tên con 4": row[28],
            "Ngày sinh con 4": row[29],
            "Tên con 5": row[30],
            "Ngày sinh con 5": row[31],
            "Ngày gửi": row[32],
            "Trạng thái": row[33],
            "Ngày cập nhật": row[34],
            "Ngày hẹn đi làm": row[35],
            "Hiệu suất": row[36],
            "Loại máy": row[37]
        })
    return result

def capnhattrangthai(sdt, trangthai):
    try:
        conn = pyodbc.connect(used_db)
        cursor = conn.cursor()
        ngaythuchien = datetime.now().date()
        query = f"UPDATE HR.Dbo.Dang_ky_thong_tin SET Trang_thai = N'{trangthai}',Ngay_cap_nhat = '{ngaythuchien}' WHERE Sdt = N'{sdt}' AND Nha_may = N'{current_user.macongty}'"
        print(query)
        cursor.execute(query)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        return False
    
def capnhatthongtinungvien(sdt, ngayhendilam, hieusuat:str, loaimay):
    try:
        conn = pyodbc.connect(used_db)
        cursor = conn.cursor()
        if not hieusuat.isdigit():
            hieusuat = 0 
        query = f"UPDATE HR.Dbo.Dang_ky_thong_tin SET Ngay_hen_di_lam = '{ngayhendilam}',Hieu_suat = '{hieusuat}',Loai_may='{loaimay}' WHERE Sdt = N'{sdt}' AND Nha_may = N'{current_user.macongty}'"
        print(query)
        cursor.execute(query)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        return False
    
def themnhanvienmoi(nhanvienmoi):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"INSERT INTO HR.Dbo.Danh_sach_CBCNV VALUES {nhanvienmoi}"
    try:
        print(query)
        cursor.execute(query)
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(e)
        return False

def xoadautrongten(s):
    s = re.sub(r'[àáạảãâầấậẩẫăằắặẳẵ]', 'a', s)
    s = re.sub(r'[ÀÁẠẢÃĂẰẮẶẲẴÂẦẤẬẨẪ]', 'A', s)
    s = re.sub(r'[èéẹẻẽêềếệểễ]', 'e', s)
    s = re.sub(r'[ÈÉẸẺẼÊỀẾỆỂỄ]', 'E', s)
    s = re.sub(r'[òóọỏõôồốộổỗơờớợởỡ]', 'o', s)
    s = re.sub(r'[ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ]', 'O', s)
    s = re.sub(r'[ìíịỉĩ]', 'i', s)
    s = re.sub(r'[ÌÍỊỈĨ]', 'I', s)
    s = re.sub(r'[ùúụủũưừứựửữ]', 'u', s)
    s = re.sub(r'[ƯỪỨỰỬỮÙÚỤỦŨ]', 'U', s)
    s = re.sub(r'[ỳýỵỷỹ]', 'y', s)
    s = re.sub(r'[ỲÝỴỶỸ]', 'Y', s)
    s = re.sub(r'[Đ]', 'D', s)
    s = re.sub(r'[đ]', 'd', s)
    return s

def themnhanvienvaomita(mst,hoten):
    conn = pyodbc.connect(mcc_db)
    cursor = conn.cursor()
    tenchamcong = xoadautrongten(hoten)
    query = f"INSERT INTO MITACOSQL.Dbo.NHANVIEN (MaNhanVien,TenNhanVien,MaChamCong,TenChamCong) VALUES ('{mst}','{tenchamcong}','{mst}','{tenchamcong}')"
    print(query)
    cursor.execute(query)
    conn.commit()
    conn.close()
    
def xoanhanvien(MST):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    try:
        query=f"DELETE FROM HR.Dbo.Danh_sach_CBCNV WHERE MST = '{MST}' AND Factory = N'{current_user.macongty}'"
        print(query)
        cursor.execute(query)
        # delete from Danh_sach_CBCNV where MST=9985 and Factory='NT1';
        conn.commit()
        conn.close()
        return f"{MST} đã xoá thành công"
    except Exception as e:
        print(e)
        return f"{MST} đã xoá thất bại"
    
def laymasothemoi():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT TOP 1 MST FROM HR.dbo.Danh_sach_CBCNV WHERE Factory = '{current_user.macongty}' ORDER BY CAST(MST AS INT) DESC"
    print(query)
    result =  cursor.execute(query).fetchone()
    conn.close()
    if result:
        return result[0]
    return 0

def laydanhsachloithe(chuyen=None, bophan=None, ngay=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    
    query = f"SELECT * FROM HR.dbo.Danh_sach_loi_the WHERE Nha_may = '{current_user.macongty}'"
    
    if chuyen:
        query += f"AND Chuyen_to = '{chuyen}' "
    if bophan:
        query += f"AND Bo_phan = '{bophan}' "
    if ngay:
        query += f"AND NgayCham = '{ngay}' "
    
    query += "ORDER BY CAST(MST AS INT) ASC, NgayCham DESC"
    print(query)
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append({
           "Nhà máy": row[0],
           "MST": row[1],
           "Họ tên": row[2],
           "Số điện thoại": row[3],
           "Chức vụ": row[4],
           "Chuyền tổ": row[5],
           "Bộ phận": row[6],
           "Ngày chấm": row[7],
           "Giờ vào": row[8],
           "Giờ ra": row[9],
           "Ca": row[10],
           "Bắt đầu": row[11],
           "Kết thúc": row[12],
           "Phân loại": row[13],
           "Số phút thiếu": row[14]
        })
    return result

def laydanhsachchuyen():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    # ngaykiemtra= (datetime.now()-timedelta(days=1)).date()
    query = f"SELECT DISTINCT Line FROM HR.dbo.Danh_sach_CBCNV WHERE Factory = '{current_user.macongty}'"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row[0])
    return result

def laydanhsachbophan():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    # ngaykiemtra= (datetime.now()-timedelta(days=1)).date()
    query = f"SELECT DISTINCT Department FROM HR.dbo.Danh_sach_CBCNV WHERE Factory = '{current_user.macongty}'"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row[0])
    return result

def laydanhsachchamcong(mst=None, tungay=None, denngay=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if mst:
        if tungay:
            if denngay: 
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE MST='{mst}' AND '{tungay}' <= Ngay AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE MST='{mst}' AND '{tungay}' <= Ngay AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
        else:
            if denngay:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE MST='{mst}' AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE MST='{mst}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    else:
        if tungay:
            if denngay: 
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE '{tungay}' <= Ngay AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE '{tungay}' <= Ngay AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
        else:
            if denngay:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    # query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result

def laydanhsachchamcongthucte(mst=None, tungay=None, denngay=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if mst:
        if tungay:
            if denngay: 
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE MST='{mst}' AND '{tungay}' <= Ngay AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE MST='{mst}' AND '{tungay}' <= Ngay AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
        else:
            if denngay:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE MST='{mst}' AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE MST='{mst}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    else:
        if tungay:
            if denngay: 
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE '{tungay}' <= Ngay AND Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE '{tungay}' <= Ngay AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
        else:
            if denngay:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE Ngay <= '{denngay}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    # query = f"SELECT * FROM HR.dbo.Bang_cham_cong_thuc_te WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay DESC, Bo_phan ASC, Chuyen_to ASC"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result

def laydanhsachdiemdanhbu(mst=None, chuyen=None, bophan=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if mst:
        query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE MST='{mst}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
    else:
        if chuyen:
            if bophan:
                query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE Line='{chuyen}' AND Bo_phan='{bophan}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE Line='{chuyen}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
        else:
            if bophan:
                query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE Bo_phan='{bophan}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
                
    # query = f"SELECT * FROM HR.dbo.Diem_danh_bu WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay_diem_danh DESC, Bo_phan ASC, Line ASC"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result

def laydanhsachxinnghiphep(mst,chuyen,bophan):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if mst:
        query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE MST='{mst}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
    else:
        if chuyen:
            if bophan:
                query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE Line='{chuyen}' AND Bo_phan='{bophan}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE Line='{chuyen}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
        else:
            if bophan:
                query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE Bo_phan='{bophan}' AND Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
            else:
                query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
    # query = f"SELECT * FROM HR.dbo.Xin_nghi_phep WHERE Nha_may = '{current_user.macongty}' ORDER BY Ngay_nghi_phep DESC, Bo_phan ASC, Line ASC"
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result

def laycacbophanduocduyet(mst,bophan):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT COUNT(*) FROM HR.dbo.Phan_quyen_phe_duyet WHERE Nha_may = '{current_user.macongty}' AND MST = '{mst}' AND Bo_phan = '{bophan}'"
    result = cursor.execute(query).fetchone()
    conn.close()
    if result[0] > 0:
        return True
    else:
        return False

def capnhat_diemdanhbu(mst,ngay,loaidiemdanh):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"UPDATE HR.dbo.Diem_danh_bu SET Trang_thai = N'Đã phê duyệt' WHERE Nha_may = '{current_user.macongty}' AND MST = '{mst}' AND Ngay_diem_danh = '{ngay}' AND Loai_diem_danh = N'{loaidiemdanh}'"
    print(query)
    cursor.execute(query)
    conn.commit()
    conn.close()

def capnhat_xinnghiphep(mst,ngay):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"UPDATE HR.dbo.Xin_nghi_phep SET Trang_thai = N'Đã phê duyệt' WHERE Nha_may = '{current_user.macongty}' AND MST = '{mst}' AND Ngay_nghi_phep = '{ngay}'"
    print(query)
    cursor.execute(query)
    conn.commit()
    conn.close()

def insert_tangca(nhamay,mst,hoten,chucvu,chuyen,phongban,ngay,giobatdau,gioketthuc):
    
    if chucvu=='nan':
        chucvu = 'Không'
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"INSERT INTO HR.dbo.Dang_ky_tang_ca VALUES (N'{nhamay}','{mst}',N'{hoten}',N'{chucvu}',N'{chuyen}',N'{phongban}','{ngay}','{giobatdau}','{gioketthuc}')"
    print(query)
    try:
        cursor.execute(query)
        conn.commit()
        conn.close()
    except Exception as e:
        print(e)
        conn.close()
        
def laydanhsachtangca(mst=None,phongban=None,ngayxem=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    
    query = f"SELECT * FROM HR.dbo.Dang_ky_tang_ca WHERE Nha_may = '{current_user.macongty}'"
    if mst:
        query += f"AND MST = '{mst}' "
    if phongban:
        query += f"AND Bo_phan = '{phongban}' "
    if ngayxem:
        query += f"AND Ngay_dang_ky = '{ngayxem}'"
    query += f" ORDER BY Ngay_dang_ky desc, CAST(MST as INT) asc"
    print(query)
    rows = cursor.execute(query).fetchall()
    # print(rows)
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result
    
def laydanhsachbaocom(chuyen=None,phongban=None,ngayxem=None):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    if chuyen:
        if phongban:
            if ngayxem:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Chuyen_to = '{chuyen}' AND Bo_phan = '{phongban}' AND Nha_may = '{current_user.macongty}' AND NgayCham = '{ngayxem}'"
            else:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Chuyen_to = '{chuyen}' AND Bo_phan = '{phongban}' AND Nha_may = '{current_user.macongty}'"
        else:
            if ngayxem:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Chuyen_to = '{chuyen}' AND Nha_may = '{current_user.macongty}' AND NgayCham = '{ngayxem}'"
            else:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Chuyen_to = '{chuyen}' AND Nha_may = '{current_user.macongty}'"
    else:
        if phongban:
            if ngayxem:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Bo_phan = '{phongban}' AND Nha_may = '{current_user.macongty}' AND NgayCham = '{ngayxem}'"
            else:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Bo_phan = '{phongban}' AND Nha_may = '{current_user.macongty}'"
        else:
            if ngayxem:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Nha_may = '{current_user.macongty}' AND NgayCham = '{ngayxem}'"
            else:
                query = f"SELECT * FROM HR.dbo.Bao_com WHERE Nha_may = '{current_user.macongty}'"
                
    print(query)
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result

def laydanhsachkyluat():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Xu_ly_ky_luat WHERE Nha_may = '{current_user.macongty}'"
    print(query)
    rows = cursor.execute(query).fetchall()
    conn.close()
    result = []
    for row in rows:
        result.append(row)
    return result    

def themdanhsachkyluat(mst,hoten,chucvu,bophan,chuyento,ngayvao,ngayvipham,diadiem,ngaylapbienban,noidung,bienphap):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"INSERT INTO HR.dbo.Xu_ly_ky_luat VALUES('{diadiem}','{mst}',N'{hoten}',N'{chucvu}','{chuyento}','{bophan}','{ngayvao}','{ngayvipham}','{ngaylapbienban}','{diadiem}',N'{noidung}',N'{bienphap}')"
    print(query)
    cursor.execute(query)
    conn.commit()
    conn.close()

def themdoicamoi(mst,cacu,camoi):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    ngayhientai = datetime.now().date()
    ngaybatdau = datetime.now().date() + timedelta(days=1)
    ngayketthuc = datetime(2054,12,31).date()
    query = f"UPDATE HR.dbo.Dang_ky_ca_lam_viec SET Den_ngay = '{ngayhientai}' WHERE MST = '{mst}' AND Factory = '{current_user.macongty}' AND Den_ngay = '{ngayketthuc}'"
    print(query)
    cursor.execute(query)
    conn.commit()
    query1 = f"INSERT INTO HR.dbo.Dang_ky_ca_lam_viec VALUES('{mst}','{current_user.macongty}','{ngaybatdau}','{ngayketthuc}','{camoi}')"
    print(query1)
    cursor.execute(query1)
    conn.commit()
    conn.close()

def laycahientai(mst):
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    ngayketthuc = datetime(2054,12,31)
    query = f"SELECT * FROM HR.dbo.Dang_ky_ca_lam_viec WHERE MST = '{mst}' AND Factory = '{current_user.macongty}' AND Den_ngay = '{ngayketthuc}'"
    print(query)
    row = cursor.execute(query).fetchone()
    if row:
        return row[-1]
    return None

def laydanhsachykienphanmem():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    ngayketthuc = datetime(2054,12,31)
    query = f"SELECT * FROM HR.dbo.Y_kien_dong_gop WHERE Nha_may = '{current_user.macongty}'"
    print(query)
    rows = cursor.execute(query).fetchall()
    result =[]
    for row in rows:
        result.append(row)
    return result 

def laydanhsachyeucautuyendung():
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"SELECT * FROM HR.dbo.Yeu_cau_tuyen_dung"
    print(query)
    rows = cursor.execute(query).fetchall()
    result =[]
    for row in rows:
        result.append(row)
    return result 

def themyeucautuyendungmoi(bophan,vitri,soluong,mota,thoigiandukien,phanloai, mucluong):

    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()
    query = f"INSERT INTO HR.dbo.Yeu_cau_tuyen_dung VALUES('{bophan}',N'{vitri}','{soluong}',N'{mota}','{thoigiandukien}',N'{phanloai}',N'{mucluong}',NULL,NULL,NULL)"
    print(query)
    try:
        cursor.execute(query)
        conn.commit()
    except Exception as e:
        print(e)
    

######################################################################################################################################################
""" ALL ROUTE"""
######################################################################################################################################################

@app.route('/unauthorized')
def unauthorized():
    return "Bạn không có quyền truy cập vui lòng chọn mục khác", 401

@app.errorhandler(404)
def page_not_found(e):
    return render_template('blank.html'), 404

@app.route('/admin', methods=["GET"])
@login_required
@roles_required('sa')
def admin_template():
    page = request.args.get('page', 1, type=int)
    per_page = 10
    mst = request.args.get('mst', None, type=str)
    if mst:
        users_paginated = [Users.query.filter_by(masothe=mst).first()]
    else:
        hoten = request.args.get('hoten', None, type=str)
        if hoten:
            search_pattern = f"%{hoten}%"
            users_paginated = Users.query.filter(Users.hoten.like(search_pattern)).paginate(page=page, per_page=per_page, error_out=False)
        else:
            users_paginated = Users.query.paginate(page=page, per_page=per_page, error_out=False)
    cacrole= ['sa','user','tuyendung','cong','luong','nhansu','developer']
    return render_template('admin.html', users=users_paginated,cacrole=cacrole)

@app.route('/register', methods=["POST"])
def register():
    if request.method == "POST":
        if request.args.get("macongty") == "NT1":
            tencty = "Công ty cổ phần sản xuất Nam Thuận"
        elif request.args.get("macongty") == "NT2":
            tencty = "Công ty cổ phần Nam Thuận Nghệ An"
        elif request.args.get("macongty") == "NT0":    
            tencty = "Công ty cổ phần tập đoàn Nam Thuận"
        user = Users(masothe=request.args.get("masothe"),
                     hoten=request.args.get("hoten"),
                     phongban=request.args.get("phongban"),
                     macongty=request.args.get("macongty"),
                     matkhau=request.args.get("matkhau"),
                     tencongty = tencty
                     )
        try:
            db.session.add(user)
            db.session.commit()
        except Exception as e:
            return str(e)
        print(user)
        
        return "OK"
 
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        print(request.form)
        user = Users.query.filter_by(
            masothe=request.form.get("masothe"),
            macongty=request.form.get("congty")).first()
        print(user)
        if not user:
            return redirect(url_for("login"))
        if user.matkhau == request.form.get("matkhau"):
            login_user(user)
            return redirect(url_for("home"))
        else:
            return redirect(url_for("login"))
    return render_template("login.html")

@app.route("/logout", methods=["GET", "POST"])
def logout():
    logout_user()
    return redirect(url_for("login"))

@app.route("/home", methods=['GET','POST'])
@login_required
def index():
    return redirect("/")

@app.route("/", methods=['GET','POST'])
@login_required
def home():
    if request.method == "POST":
        mst = request.form.get("mst")
        hoten = request.form.get("hoten")
        sdt = request.form.get("sdt")
        cccd = request.form.get("cccd")
        gioitinh = request.form.get("gioitinh")
        ngayvao = request.form.get("ngayvao")
        ngaynghi = request.form.get("ngaynghi")
        ngaykyhd = request.form.get("ngaykyhd")
        ngayhethanhd = request.form.get("ngayhethanhd")
        phongban = request.form.get("phongban")
        trangthai = request.form.get("trangthai")
        hccategory = request.form.get("hccategory")
        return redirect(url_for('home', 
                                mst=mst, 
                                hoten=hoten, 
                                sdt=sdt, 
                                cccd=cccd, 
                                gioitinh=gioitinh, 
                                ngayvao=ngayvao, 
                                ngaynghi=ngaynghi, 
                                ngaykyhd=ngaykyhd, 
                                ngayhethanhd=ngayhethanhd, 
                                phongban=phongban,
                                trangthai=trangthai,
                                hccategory=hccategory
                                ))
            
    elif request.method == "GET":
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        sdt = request.args.get("sdt")
        cccd = request.args.get("cccd")
        gioitinh = request.args.get("gioitinh")
        ngayvao = request.args.get("ngayvao")
        ngaynghi = request.args.get("ngaynghi")
        ngaykyhd = request.args.get("ngaykyhd")
        ngayhethanhd = request.args.get("ngayhethanhd")
        phongban = request.args.get("phongban")
        trangthai = request.args.get("trangthai")
        hccategory = request.args.get("hccategory")
        
        users = laydanhsachuser(mst, hoten, sdt, cccd, gioitinh, ngayvao, ngaynghi, ngaykyhd, ngayhethanhd, phongban, trangthai, hccategory)   
        count = len(users)
        cacphongban = laycacphongban()
        cacto = laycacto()
        cactrangthai = laycactrangthai()
        cachccategory = laycachccategory()
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(users)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_users = users[start:end]
        
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        
        return render_template("home.html", users=paginated_users, 
                            cacphongban=cacphongban, cacto=cacto,
                            page="Trang chủ", pagination=pagination,
                            cactrangthai=cactrangthai,count=count,
                            cachccategory=cachccategory)

@app.route("/muc2_1", methods=["GET","POST"])
@login_required
@roles_required('tuyendung','developer','sa','nhansu')
def danhsachdangkytuyendung():
    if request.method == "GET":
        sdt = request.args.get("sdt")
        cccd = request.args.get("cccd")
        users = laydanhsachdangkytuyendung(sdt,cccd)
        return render_template("2_1.html", page="2.1 Danh sách đăng ký tuyển dụng", users=users)
    if request.method == "POST":
        sdt = request.form.get("sdt")
        ngayhendilam = request.form.get("ngayhendilam")
        hieusuat = request.form.get("hieusuat")
        loaimay = request.form.get("loaimay")
        capnhatthongtinungvien(sdt,ngayhendilam,hieusuat,loaimay)
        return redirect(f"muc2_1?sdt={sdt}")

@app.route("/muc2_2_1", methods=["GET","POST"])
@login_required
@roles_required('tbp','developer','sa','nhansu')
def dangkytuyendung():
    if request.method == "GET":
        danhsach = laydanhsachyeucautuyendung()
        return render_template("2_2_1.html", page="2.2.1 Thêm yêu cầu tuyển dụng",danhsach=danhsach)
    elif request.method == "POST":
        bophan = request.form.get("bophan")
        vitri = request.form.get("vitri")
        soluong = request.form.get("soluong")
        mota = request.form.get("mota")
        thoigiandukien = request.form.get("thoigiandukien")
        phanloai = request.form.get("phanloai")
        mucluongtu = request.form.get("mucluongtu")
        mucluongden = request.form.get("mucluongden")
        mucluong = f"{mucluongtu} - {mucluongden} triệu VNĐ"
        themyeucautuyendungmoi(bophan,vitri,soluong,mota,thoigiandukien,phanloai,mucluong)
        return redirect("muc2_2_1")
    
@app.route("/muc2_2_2", methods=["GET","POST"])
@login_required
@roles_required('tbp','developer','sa','nhansu')
def pheduyettuyendung():   
    if request.method == "GET":
        danhsach = laydanhsachyeucautuyendung()
        return render_template("2_2_2.html", page="2.2.2 Phê duyệt yêu cầu tuyển dụng",danhsach=danhsach)
    
    
@app.route("/muc3_1", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','deveploper')
def nhapthongtinlaodongmoi():
    
    if request.method == "GET":
        data= request.args.get("data")
        masothe = int(laymasothemoi())+1
        cacvitri= laycacvitri()
        cacto = laycacto()
        cacca = laycacca()
        data=request.args.get("scan-qrcode")
        datenow = datetime.now()
        macongty = current_user.macongty 
        return render_template("3_1.html", 
                                page="3.1 Nhập thông tin lao động mới",
                                qrcccd=data,
                                masothe=masothe,
                                ngaybatdau=datenow,
                                cacvitri=cacvitri,
                                cacto=cacto,
                                cacca=cacca,
                                macongty=macongty)
    elif request.method == "POST":

        anh = f"N'{request.form.get("anh")}'"
        masothe = f"'{request.form.get("masothe")}'"
        thechamcong = "NULL"
        hoten = f"N'{request.form.get("hoten")}'"
        ngaysinh = f"'{datetime.strptime(request.form.get("ngaysinh"),"%d/%m/%Y").date()}'" if request.form.get("ngaysinh") else "NULL"
        gioitinh = f"N'{request.form.get("gioitinh")}'"
        cmt = f"'{request.form.get("cmt")}'"
        cccd = f"'{request.form.get("cccd")}'"
        ngaycapcccd = f"'{datetime.strptime(request.form.get("ngaycap"),"%d/%m/%Y").date()}'" if request.form.get("ngaycap") else "NULL"
        thuongtru = f"N'{request.form.get("thuongtru")}'" if request.form.get("thuongtru") else "NULL"
        noisinh = f"N'{request.form.get("noisinh")}'" if request.form.get("noisinh") else "NULL"
        tamtru = f"N'{request.form.get("tamtru")}'" if request.form.get("tamtru") else "NULL"
        quoctich = f"N'{request.form.get("quoctich")}'" if request.form.get("quoctich") else "NULL"
        dantoc = f"N'{request.form.get("dantoc")}'" if request.form.get("dantoc") else "NULL"
        tongiao = f"N'{request.form.get("tongiao")}'" if request.form.get("tongiao") else "NULL"
        hocvan = f"N'{request.form.get("hocvan")}'" if request.form.get("hocvan") else "NULL"
        thonxom = f"N'{request.form.get("thonxom")}'" if request.form.get("thonxom") else "NULL"
        phuongxa = f"N'{request.form.get("phuongxa")}'" if request.form.get("phuongxa") else "NULL"
        quanhuyen = f"N'{request.form.get("quanhuyen")}'" if request.form.get("quanhuyen") else "NULL"
        tinhthanhpho = f"N'{request.form.get("tinhthanhpho")}'" if request.form.get("tinhthanhpho") else "NULL"
        nganhang = f"N'{request.form.get("nganhang")}'" if request.form.get("nganhang") else "NULL"
        sotaikhoan = f"'{request.form.get("sotaikhoan")}'" if request.form.get("sotaikhoan") else "NULL"
        dienthoai = f"'{request.form.get("dienthoai")}'" if request.form.get("dienthoai") else "NULL"
        sobhxh = f"'{request.form.get("sobhxh")}'" if request.form.get("sobhxh") else 'NULL'
        masothue = f"'{request.form.get("masothue")}'" if request.form.get("masothue") else 'NULL'
        connho = f"N'{request.form.get("connho")}'" if request.form.get("connho") else 'NULL'
        tencon1 = f"N'{request.form.get("tenconmnho1")}'" if request.form.get("tenconmnho1") else 'NULL'
        ngaysinhcon1 = f"'{datetime.strptime(request.form.get("ngaysinhcon1"),"%Y-%m-%d").date()}'" if request.form.get("ngaysinhcon1") else 'NULL'
        tencon2 = f"N'{request.form.get("tenconmnho2")}'" if request.form.get("tenconmnho2") else 'NULL'
        ngaysinhcon2 = f"'{datetime.strptime(request.form.get("ngaysinhcon2"),"%Y-%m-%d").date()}'" if request.form.get("ngaysinhcon2") else 'NULL'
        tencon3 = f"N'{request.form.get("tenconmnho3")}'" if request.form.get("tenconmnho3") else 'NULL'
        ngaysinhcon3 = f"'{datetime.strptime(request.form.get("ngaysinhcon3"),"%Y-%m-%d").date()}'" if request.form.get("ngaysinhcon3") else 'NULL'
        tencon4 = f"N'{request.form.get("tenconmnho4")}'" if request.form.get("tenconmnho4") else 'NULL'
        ngaysinhcon4 = f"'{datetime.strptime(request.form.get("ngaysinhcon4"),"%Y-%m-%d").date()}'" if request.form.get("ngaysinhcon4") else 'NULL'
        tencon5 = f"N'{request.form.get("tenconmnho5")}'" if request.form.get("tenconmnho5") else 'NULL'
        ngaysinhcon5 = f"'{datetime.strptime(request.form.get("ngaysinhcon5"),"%Y-%m-%d").date()}'" if request.form.get("ngaysinhcon5") else 'NULL'
        jobdetailvn = f"N'{request.form.get("vitri")}'"
        line = f"'{request.form.get("line")}'"
        calamviec = request.form.get("calamviec")
        factory = f"'{current_user.macongty}'"
        hccategory = f"N'{request.form.get("hccategory")}'"
        gradecode = f"N'{request.form.get("gradecode")}'"
        department = f"N'{request.form.get("phongban")}'"
        chucvu = f"N'{request.form.get("chucvu")}'"
        employeetype = f"N'{request.form.get("loailaodong")}'"
        sectioncode = f"N'{request.form.get("mabophan")}'"
        sectiondescription = f"N'{request.form.get("bophan")}'"
        jobdetailvn = f"N'{request.form.get("vitrien")}'"
        positioncode = f"N'{request.form.get("mavitri")}'"
        positioncodedescription = f"N'{request.form.get("tenvitri")}'"
        nguoithan = f"N'{request.form.get("nguoithan")}'" if request.form.get("nguoithan") else 'NULL'
        sdtnguoithan = f"N'{request.form.get("sodienthoainguoithan")}'" if request.form.get("sodienthoainguoithan") else 'NULL'
        luongcoban = f"'{request.form.get("luongcoban")}'" if request.form.get("luongcoban") else 'NULL'
        tongphucap = f"'{request.form.get("tongphucap")}'" if request.form.get("tongphucap") else 'NULL'
        kieuhopdong = request.form.get("kieuhopdong")
        print(kieuhopdong)
        if kieuhopdong == "HĐ thử việc":
            kieuhopdong = "N'HĐ thử việc'"
            ngaybatdauthuviec = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayvao = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuc = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngayketthucthuviec = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngaybatdauhdcthl1 = "NULL"
            ngayketthuchdcthl1 = "NULL"
            ngaybatdauhdcthl2 = "NULL"
            ngayketthuchdcthl2 = "NULL"
            ngaybatdauhdvth = "NULL"
        elif kieuhopdong == "HĐ có thời hạn lần 1":
            kieuhopdong = "N'HĐ có thời hạn lần 1'"
            ngayvao = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuc = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngaybatdauhdcthl1 = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuchdcthl1 = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngaybatdauthuviec = "NULL"
            ngayketthucthuviec = "NULL"
            ngaybatdauhdcthl2 = "NULL"
            ngayketthuchdcthl2 = "NULL"
            ngaybatdauhdvth = "NULL"
        elif kieuhopdong == "HĐ có thời hạn lần 2":
            kieuhopdong = "N'HĐ có thời hạn lần 2'"
            ngayvao = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuc = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngaybatdauhdcthl2 = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuchdcthl2 = f"'{request.form.get("ngayKetThuc")}'" if request.form.get("ngayKetThuc") else 'NULL'
            ngaybatdauthuviec = "NULL"
            ngayketthucthuviec = "NULL"
            ngaybatdauhdcthl1 = "NULL"
            ngayketthuchdcthl1 = "NULL"
            ngaybatdauhdvth = "NULL"
        elif kieuhopdong == "HĐ vô thời hạn":
            kieuhopdong = "N'HĐ vô thời hạn'"
            ngayvao = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngayketthuc = 'NULL'
            ngaybatdauhdvth = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            ngaybatdauthuviec = "NULL"
            ngayketthucthuviec = "NULL"
            ngaybatdauhdcthl1 = "NULL"
            ngayketthuchdcthl1 = "NULL"
            ngaybatdauhdcthl2 = "NULL"
            ngayketthuchdcthl2 = "NULL"
        else:
            ngayvao = f"'{request.form.get("ngayBatDau")}'" if request.form.get("ngayBatDau") else 'NULL'
            kieuhopdong = "NULL"
            ngayketthuc = 'NULL'
            ngaybatdauthuviec = "NULL"
            ngayketthucthuviec = "NULL"
            ngaybatdauhdcthl1 = "NULL"
            ngayketthuchdcthl1 = "NULL"
            ngaybatdauhdcthl2 = "NULL"
            ngayketthuchdcthl2 = "NULL"
            ngaybatdauhdvth = "NULL"
        nhanvienmoi = f"({masothe},{thechamcong},{hoten},{dienthoai},{ngaysinh},{gioitinh},{cccd},{ngaycapcccd},N'Cục cảnh sát',{cmt},{thuongtru},{thonxom},{phuongxa},{quanhuyen},{tinhthanhpho},{dantoc},{quoctich},{tongiao},{hocvan},{noisinh},{tamtru},{sobhxh},{masothue},{nganhang},{sotaikhoan},{connho},{tencon1},{ngaysinhcon1},{tencon2},{ngaysinhcon2},{tencon3},{ngaysinhcon3},{tencon4},{ngaysinhcon4},{tencon5},{ngaysinhcon5},{anh},{nguoithan}, {sdtnguoithan},{kieuhopdong},{ngayvao},{ngayketthuc},{jobdetailvn},{hccategory},{gradecode},{factory},{department},{chucvu},{sectioncode},{sectiondescription},{line},{employeetype},{jobdetailvn},{positioncode},{positioncodedescription},{luongcoban},N'Không',{tongphucap},{ngayvao},NULL,N'Đang làm việc',{ngayvao},'1',{ngaybatdauthuviec},{ngayketthucthuviec},{ngaybatdauhdcthl1},{ngayketthuchdcthl1},{ngaybatdauhdcthl2},{ngayketthuchdcthl2},{ngaybatdauhdvth},'N')"             
        if themnhanvienmoi(nhanvienmoi):
            # themnhanvienvaomita(request.form.get("masothe"),request.form.get("hoten"))
            themdoicamoi(request.form.get("masothe"),None,calamviec)
            return redirect("/muc3_1")
        else:
            masothe = int(laymasothemoi())+1
            cacvitri= laycacvitri()
            cacto = laycacto()
            cacca = laycacca()
            return redirect("/muc3_1")
        
@app.route("/muc3_2", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','deveploper')
def thaydoithongtinlaodong():
    
    if request.method == "GET":
        return render_template("3_2.html", page="3.2 Thay đổi thông tin người lao động")
    else:
        thechamcong = request.form.get("thechamcong")
        cccd = request.form.get("cccd")
        ngaycapcccd = request.form.get("ngaycapcccd")
        hoten = request.form.get("hoten")
        ngaysinh = request.form.get("ngaysinh")
        gioitinh = request.form.get("gioitinh")
        cmt = request.form.get("cmt")
        quoctich = request.form.get("quoctich")
        dienthoai = request.form.get("dienthoai")
        thonxom = request.form.get("thonxom")
        phuongxa = request.form.get("phuongxa")
        quanhuyen = request.form.get("quanhuyen")
        tinhthanhpho = request.form.get("tinhthanhpho")
        dantoc = request.form.get("dantoc")
        tongiao = request.form.get("tongiao")
        hocvan = request.form.get("hocvan")
        masothue = request.form.get("masothue")
        nganhang = request.form.get("nganhang")
        sotaikhoan = request.form.get("sotaikhoan")
        connho = request.form.get("connho")
        mst = request.form.get("mst")
        tenconnho1 = request.form.get("tenconnho1")
        tenconnho2 = request.form.get("tenconnho2")
        tenconnho3 = request.form.get("tenconnho3")
        tenconnho4 = request.form.get("tenconnho4")
        tenconnho5 = request.form.get("tenconnho5")
        ngaysinhcon1 = request.form.get("ngaysinhcon1")
        ngaysinhcon2 = request.form.get("ngaysinhcon2")
        ngaysinhcon3 = request.form.get("ngaysinhcon3")
        ngaysinhcon4 = request.form.get("ngaysinhcon4")
        ngaysinhcon5 = request.form.get("ngaysinhcon5")
        jobtitlevn = request.form.get("jobtitlevn") 
        jobtitleen = request.form.get("jobtitleen") 
        positioncode = request.form.get("positioncode")
        positioncodedescription = request.form.get("positioncodedescription")
        chucvu =  request.form.get("chucvu")
        line = request.form.get("line")
        department = request.form.get("department")
        sectioncode = request.form.get("sectioncode")
        sectiondescription = request.form.get("sectiondescription")
        hccategory = request.form.get("hccategory")
        employeetype = request.form.get("employeetype")
        gradecode = request.form.get("gradecode")
        factory = request.form.get("factory")
        kieuhopdong = request.form.get("kieuhopdong")
        ngaybatdau = request.form.get("ngaybatdau")
        ngayketthuc = request.form.get("ngayketthuc")
        mucluong = request.form.get("mucluong")
        phucap = request.form.get("phucap")
        tienphucap = request.form.get("tienphucap")
        
        query = f"UPDATE HR.dbo.Danh_sach_CBCNV SET "
        if thechamcong:
            query += f"The_cham_cong = '{thechamcong}',"
        else:
            query += f"The_cham_cong = NULL,"
            
        if cccd: 
            query += f"CCCD = '{cccd}',"
        else:
            query += f"CCCD = NULL,"
            
        if ngaycapcccd: 
            query += f"Ngay_cap = '{ngaycapcccd}',"
        else:
            query += f"Ngay_cap = NULL,"
            
        if hoten: 
            query += f"Ho_ten = N'{hoten}',"
        else:
            query += f"Ho_ten = NULL,"
            
        if ngaysinh: 
            query += f"Ngay_sinh = '{ngaysinh}',"
        else:
            query += f"Ngay_sinh = NULL,"
            
        if gioitinh: 
            query += f"Gioi_tinh = N'{gioitinh}',"
        else:
            query += f"Gioi_tinh = NULL,"
                
        if cmt: 
            query += f"CMT = N'{cmt}',"
        else:
            query += f"CMT = NULL,"   
            
        if quoctich: 
            query += f"Quoc_tich = N'{quoctich}',"
        else:
            query += f"Quoc_tich = NULL," 
            
        if dienthoai: 
            query += f"Sdt = N'{dienthoai}',"
        else:
            query += f"Sdt = NULL,"
            
        if thonxom: 
            query += f"Thon_xom = N'{thonxom}',"
        else:
            query += f"Thon_xom = NULL,"
            
        if phuongxa: 
            query += f"Phuong_xa = N'{phuongxa}',"
            
        if quanhuyen: 
            query += f"Quan_huyen = N'{quanhuyen}',"
        else:
            query += f"Quan_huyen = NULL,"
            
        if tinhthanhpho: 
            query += f"Tinh_TP = N'{tinhthanhpho}',"
        else:
            query += f"Tinh_TP = NULL,"
            
        if dantoc: 
            query += f"Dan_toc = N'{dantoc}',"
        else:
            query += f"Dan_toc = NULL,"
                
        if tongiao: 
            query += f"Ton_giao = N'{tongiao}',"
        else:
            query += f"Ton_giao = NULL,"
                
        if hocvan: 
            query += f"Trinh_do = N'{hocvan}',"
        else:
            query += f"Trinh_do = NULL,"
            
        if masothue: 
            query += f"Ma_so_thue = N'{masothue}',"
        else:
            query += f"Ma_so_thue = NULL,"   
            
        if nganhang: 
            query += f"Ngan_hang = N'{nganhang}',"
        else:
            query += f"Ngan_hang = NULL,"   
            
        if sotaikhoan: 
            query += f"So_tai_khoan = N'{sotaikhoan}',"
        else:
            query += f"So_tai_khoan = NULL,"
            
        if connho: 
            query += f"Con_nho = N'{connho}',"
        else:
            query += f"Con_nho = NULL,"

        if tenconnho1: 
            query += f"Ten_con_nho_1 = N'{tenconnho1}',"
        else:
            query += f"Ten_con_nho_1 = NULL,"
        
        if tenconnho2: 
            query += f"Ten_con_nho_2 = N'{tenconnho2}',"
        else:
            query += f"Ten_con_nho_2 = NULL,"
            
        if tenconnho3: 
            query += f"Ten_con_nho_3 = N'{tenconnho3}',"
        else:
            query += f"Ten_con_nho_3 = NULL,"
            
        if tenconnho4: 
            query += f"Ten_con_nho_4 = N'{tenconnho4}',"
        else:
            query += f"Ten_con_nho_4 = NULL,"
        
        if tenconnho5: 
            query += f"Ten_con_nho_5 = N'{tenconnho5}',"
        else:
            query += f"Ten_con_nho_5 = NULL,"
        
        if ngaysinhcon1: 
            query += f"Ngay_sinh_con_nho_1 = '{ngaysinhcon1}',"
        else:
            query += f"Ngay_sinh_con_nho_1 = NULL,"
        
        if ngaysinhcon2: 
            query += f"Ngay_sinh_con_nho_2 = '{ngaysinhcon2}',"
        else:
            query += f"Ngay_sinh_con_nho_2 = NULL,"
            
        if ngaysinhcon3: 
            query += f"Ngay_sinh_con_nho_3 = '{ngaysinhcon3}',"
        else:
            query += f"Ngay_sinh_con_nho_3 = NULL,"
            
        if ngaysinhcon4: 
            query += f"Ngay_sinh_con_nho_4 = '{ngaysinhcon4}',"
        else:
            query += f"Ngay_sinh_con_nho_4 = NULL,"
            
        if ngaysinhcon5: 
            query += f"Ngay_sinh_con_nho_5 = '{ngaysinhcon5}',"
        else:
            query += f"Ngay_sinh_con_nho_5 = NULL,"
        
        if jobtitlevn: 
            query += f"Job_title_VN = N'{jobtitlevn}',"
        else:
            query += f"Job_title_VN = NULL,"
            
        if jobtitleen: 
            query += f"Job_title_EN = '{jobtitleen}',"
        else:
            query += f"Job_title_EN = NULL,"
            
        if positioncode: 
            query += f"Position_code = '{positioncode}',"
        else:
            query += f"Position_code = NULL,"
        
        if positioncodedescription: 
            query += f"Position_code_description = '{positioncodedescription}',"
        else:
            query += f"Position_code_description = NULL,"  
        
        if chucvu: 
            query += f"Chuc_vu = '{chucvu}',"
        else:
            query += f"Chuc_vu = NULL," 
            
        if line: 
            query += f"Line = '{line}',"
        else:
            query += f"Line = NULL,"     
        
        if department: 
            query += f"Department = '{department}',"
        else:
            query += f"Department = NULL,"  
            
        if sectioncode: 
            query += f"Section_code = '{sectioncode}',"
        else:
            query += f"Section_code = NULL,"     
        
        if sectiondescription: 
            query += f"Section_description = '{sectiondescription}',"
        else:
            query += f"Section_description = NULL," 
            
        if hccategory: 
            query += f"Headcount_category = '{hccategory}',"
        else:
            query += f"Headcount_category = NULL," 
            
        if employeetype: 
            query += f"Emp_type = '{employeetype}',"
        else:
            query += f"Emp_type = NULL," 
            
        if gradecode: 
            query += f"Grade_code = '{gradecode}',"
        else:
            query += f"Grade_code = NULL," 
            
        if factory: 
            query += f"Factory = '{factory}',"
        else:
            query += f"Factory = NULL,"
            
        if kieuhopdong:
            query += f"Loai_hop_dong = N'{kieuhopdong}',"
        else:
            query += f"Loai_hop_dong = NULL,"
            
        if ngaybatdau:
            query += f"Ngay_ky_HD = '{ngaybatdau}',"
        else:
            query += f"Ngay_ky_HD = NULL,"
          
        if ngayketthuc:
            query += f"Ngay_het_han_HD = '{ngayketthuc}',"
        else:
            query += f"Ngay_het_han_HD = NULL,"
            
        if mucluong:
            query += f"Luong_co_ban = '{mucluong}',"
        else:
            query += f"Luong_co_ban = NULL,"
            
        if phucap:
            query += f"Phu_cap = N'{phucap}',"
        else:
            query += f"Phu_cap = NULL,"
            
        if tienphucap:
            query += f"Tong_phu_cap = N'{tienphucap}',"
        else:
            query += f"Tong_phu_cap = NULL,"
               
        query = query[:-1] + f" WHERE MST = '{mst}' AND Factory='{current_user.macongty}'"
        conn = pyodbc.connect(used_db)
        cursor = conn.cursor()
        print(query)
        cursor.execute(query)
        conn.commit()
        return redirect("/muc3_2")
    
@app.route("/muc3_3", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def inhopdonglaodong():
    if request.method == "GET":
        return render_template("3_3.html", page="3.3 In hợp đồng lao động")
    elif request.method == "POST":
        bophan = request.form.get("bophan")
        kieuhopdong = request.form.get("kieuhopdong")
        mst = request.form.get("mst")
        ngaylamhopdong = request.form.get("ngaybatdau")[-2:]
        thanglamhopdong = request.form.get("ngaybatdau")[5:7]
        namlamhopdong = request.form.get("ngaybatdau")[:4]
        ngayketthuchopdong = request.form.get("ngayketthuc")[-2:]
        thangketthuchopdong = request.form.get("ngayketthuc")[5:7]
        namketthuchopdong = request.form.get("ngayketthuc")[:4]
        tennhanvien = request.form.get("hoten")
        ngaysinh = request.form.get("ngaysinh")
        gioitinh = request.form.get("gioitinh")
        thuongtru = request.form.get("thuongtru")
        cccd = request.form.get("cccd")
        ngaycapcccd = request.form.get("ngaycapcccd")
        mucluong = request.form.get("luongcoban")
        chucvu = request.form.get("chucvu")
        try:
            file = thaydoithongtinhopdong(kieuhopdong,
                                            mst,
                                            ngaylamhopdong,
                                            thanglamhopdong,
                                            namlamhopdong,
                                            ngayketthuchopdong,
                                            thangketthuchopdong,
                                            namketthuchopdong,
                                            tennhanvien,
                                            ngaysinh,
                                            gioitinh,
                                            thuongtru,
                                            cccd,
                                            ngaycapcccd,
                                            mucluong,
                                            chucvu,
                                            bophan)
            print(f"File: {file}")
            if file:
                return send_file(file, as_attachment=True, download_name="hopdonglaodong.xlsx")
            else:
                return redirect("/muc3_3")
        except:
            return redirect("/muc3_3")   

@app.route("/muc6_1", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def dieuchuyen():
    
    if request.method == "POST":
        mst = request.form["mst"]
        loaidieuchuyen = request.form["loaidieuchuyen"]
        ngaydieuchuyen = request.form.get("ngaydieuchuyen")
        
        vitricu = request.form.get("vitricu")
        vitrimoi = request.form.get("vitrimoi")
        
        vitriencu = request.form.get("vitriencu")
        vitrienmoi = request.form.get("vitrienmoi")
        
        chuyencu = request.form.get("chuyencu")
        chuyenmoi = request.form.get("chuyenmoi")
        
        gradecodecu = request.form.get("gradecodecu")
        gradecodemoi = request.form.get("gradecodemoi")
        
        sectioncodecu = request.form.get("sectioncodecu")
        sectioncodemoi = request.form.get("sectioncodemoi")
        
        hccategorycu = request.form.get("hccategorycu")
        hccategorymoi = request.form.get("hccategorymoi")
        
        departmentcu = request.form.get("departmentcu")
        departmentmoi = request.form.get("departmentmoi")
        
        sectiondescriptioncu = request.form.get("sectiondescriptioncu")
        sectiondescriptionmoi = request.form.get("sectiondescriptionmoi")
        
        employeetypecu = request.form.get("employeetypecu") 
        employeetypemoi = request.form.get("employeetypemoi")
        
        positioncodecu = request.form.get("positioncodecu") 
        positioncodemoi = request.form.get("positioncodemoi") 
        
        positioncodedescriptioncu = request.form.get("positioncodedescriptioncu") 
        positioncodedescriptionmoi = request.form.get("positioncodedescriptionmoi") 
        
        if loaidieuchuyen == "Chuyển vị trí":
            try:
                dieuchuyennhansu(mst,
                                loaidieuchuyen,
                                vitricu,
                                vitrimoi,
                                chuyencu,
                                chuyenmoi,
                                gradecodecu,
                                gradecodemoi,
                                sectioncodecu,
                                sectioncodemoi,
                                hccategorycu,
                                hccategorymoi,
                                departmentcu,
                                departmentmoi,
                                sectiondescriptioncu,
                                sectiondescriptionmoi,
                                employeetypecu,
                                employeetypemoi,
                                positioncodedescriptioncu,
                                positioncodedescriptionmoi,
                                positioncodecu,
                                positioncodemoi,
                                vitriencu,
                                vitrienmoi,
                                ngaydieuchuyen
                                )
            except Exception as e:
                print(e)
                return redirect(f"/muc6_2?mst={mst}")
            
        else:
            try:
                dichuyennghiviec(mst,
                            loaidieuchuyen,
                            vitricu,
                            vitrimoi,
                            chuyencu,
                            chuyenmoi,
                            gradecodecu,
                            gradecodemoi,
                            sectioncodecu,
                            sectioncodemoi,
                            hccategorycu,
                            hccategorymoi,
                            departmentcu,
                            departmentmoi,
                            sectiondescriptioncu,
                            sectiondescriptionmoi,
                            employeetypecu,
                            employeetypemoi,
                            positioncodedescriptioncu,
                            positioncodedescriptionmoi,
                            positioncodecu,
                            positioncodemoi,
                            vitriencu,
                            vitrienmoi,
                            ngaydieuchuyen
                            )
            except Exception as e:
                print(e)
                return redirect(f"/muc6_2?mst={mst}")
        return redirect(f"/muc6_2?mst={mst}")
    else:  
        cacvitri= laycacvitri()
        return render_template("6_1.html",
                            cacvitri=cacvitri,
                            page="6.1 Điều chuyển chức vụ, bộ phận")
    
@app.route("/muc6_2", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def lichsucongtac():
    
    if request.method == "GET":
        mst = request.args.get("mst")
        if not mst:
            mst = request.form.get("mst")
        print(mst)
        cacphongban = laycacphongban()
        danhsach = laylichsucongtactheomst(mst)
    return render_template("6_2.html", page="6.2 Lịch sử công tác",danhsach=danhsach, mst=mst, cacphongban=cacphongban)

    
@app.route("/muc7_1_1", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def khaibaochamcong():
    if request.method == "GET":
        danhsachphongban = laycacphongban()
        danhsachca = laycacca()
        cacchuyen = laycacto()
        mst = request.args.get("mst")
        calamviec = laycahientai(mst)
        return render_template("7_1_1.html",
                                page="7.1.1 Đổi ca làm việc",
                                danhsachphongban=danhsachphongban,
                                danhsachca = danhsachca,
                                cacchuyen=cacchuyen,
                                calamviec=calamviec)
    elif request.method == "POST":
        mst = request.form.get('mst')
        return redirect(f"/muc7_1_1?mst={mst}")
    
@app.route("/muc7_1_2", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def chamcongtudong():
    
    mst = request.args.get("masothe")
    tungay = request.args.get("tungay")
    denngay = request.args.get("denngay")
    if not tungay and not denngay:
        tungay = datetime.now().date()
        denngay = datetime.now().date()
    rows = laydanhsachchamcong(mst,tungay,denngay)
    return render_template("7_1_2.html", page="7.1.2 Chấm công tự động",danhsach=rows)

@app.route("/muc7_1_3", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def loichamcong():
    
    chuyen = request.args.get("chuyen")
    bophan = request.args.get("bophan")
    ngay = request.args.get("ngay")
    danh_sach_chuyen = laydanhsachchuyen()
    danh_sach_bophan = laydanhsachbophan()
    danhsachloichamcong = laydanhsachloithe(chuyen,bophan,ngay)
    return render_template("7_1_3.html",
                            page="7.1.3 Danh sách lỗi chấm công",
                            danhsachloichamcong=danhsachloichamcong,
                            danh_sach_chuyen=danh_sach_chuyen,
                            danh_sach_bophan=danh_sach_bophan)

@app.route("/muc7_1_4", methods=["GET"])
@login_required
@roles_required('cong','sa','developer')
def diemdanhbu():
    
    mst = request.args.get("mst")
    chuyen = request.args.get("chuyen")
    bophan = request.args.get("bophan")
    danh_sach_chuyen = laydanhsachchuyen()
    danh_sach_bophan = laydanhsachbophan()
    danhsach_diemdanh_bu = laydanhsachdiemdanhbu(mst,chuyen,bophan)
    return render_template("7_1_4.html",
                        page="7.1.4 Danh sách điểm danh bù",
                        danhsach_diemdanh_bu=danhsach_diemdanh_bu,
                        danh_sach_chuyen=danh_sach_chuyen,
                        danh_sach_bophan=danh_sach_bophan)
 
@app.route("/muc7_1_5", methods=["GET"])
@login_required
@roles_required('cong','sa','developer')
def xinnghiphep():
    
        mst = request.args.get("mst")
        chuyen = request.args.get("chuyen")
        bophan = request.args.get("bophan")
        danh_sach_chuyen = laydanhsachchuyen()
        danh_sach_bophan = laydanhsachbophan()
        danhsach_xinnghiphep = laydanhsachxinnghiphep(mst,chuyen,bophan)
        return render_template("7_1_5.html",
                            page="7.1.5 Danh sách xin nghỉ phép",
                            danhsach_xinnghiphep=danhsach_xinnghiphep,
                            danh_sach_chuyen=danh_sach_chuyen,
                            danh_sach_bophan=danh_sach_bophan)

@app.route("/muc7_1_6", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def dangkytangca():
    
    if request.method == "GET":
        mst = request.args.get("mst")
        phongban = request.args.get("phongban")
        ngayxem = request.args.get("ngay")
        danhsach = laydanhsachtangca(mst,phongban,ngayxem)
        return render_template("7_1_6.html", page="7.1.6 Đăng ký tăng ca",danhsach=danhsach)
    elif request.method == "POST":
        ngayxem = request.form.get("ngay")
        data = []
        for row in danhsach:
            data.append({
                "Nhà máy": row[0],
                "MST": row[1],
                "Họ tên": row[2],
                "Chức vụ": row[3],
                "Chuyền tổ": row[4], 
                "Phòng ban": row[5],
                "Ngày đăng ký": row[6],
                "Giờ tăng ca": row[7],
                "Giờ tăng ca thực tế": row[8]
            })
        df = pd.DataFrame(data)
        df.to_excel("tangca.xlsx", index=False)

        return send_file("tangca.xlsx", as_attachment=True)
        
@app.route("/muc7_1_7", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def baocom():
    
    if request.method == "GET":
        chuyen = request.args.get("chuyen")
        phongban = request.args.get("phongban")
        ngayxem = request.args.get("ngay")
        if not ngayxem:
            ngayxem = datetime.now().date()
        danhsach = laydanhsachbaocom(chuyen,phongban,ngayxem)
        count = len(danhsach)
        danhsachchuyen = laycacto()
        danhsachphongban = laycacphongban()
        return render_template("7_1_7.html", 
                            page="7.1.7 Danh sách báo cơm",
                            danhsach=danhsach,
                            count=count,
                            danhsachchuyen=danhsachchuyen,
                            danhsachphongban=danhsachphongban)
    elif request.method == "POST":
        chuyen = request.form.get("chuyen")
        phongban = request.form.get("phongban")
        ngayxem = request.form.get("ngay")
        danhsach = laydanhsachbaocom(chuyen,phongban,ngayxem)
        data = []
        for row in danhsach:
            data.append({
                "Nhà máy": row[0],
                "MST": row[1],
                "Họ tên": row[2],
                "Chức vụ": row[3],
                "Chuyền tổ": row[4], 
                "Phòng ban": row[5],
                "Ngày chấm": row[6],
                "Giờ vào": row[7],
            })
        df = pd.DataFrame(data)
        df.to_excel("baocom.xlsx", index=False)

        return send_file("baocom.xlsx", as_attachment=True)
    
@app.route("/muc7_1_8", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def baocaotonghop():
    
    return render_template("7_1_8.html", page="7.1.8 Báo cáo tổng hợp")

    
@app.route("/muc7_1_9", methods=["GET","POST"])
@login_required
@roles_required('cong','sa','developer')
def danhsachxinnghikhac():
    
    return render_template("7_1_9.html", page="7.1.9 Danh sách xin nghỉ khác")

@app.route("/muc7_2", methods=["GET","POST"])
@login_required
@roles_required('sa','developer')
def chamcongtudongthucte():
    
    mst = request.args.get("masothe")
    tungay = request.args.get("tungay")
    denngay = request.args.get("denngay")
    rows = laydanhsachchamcongthucte(mst,tungay,denngay)
    return render_template("7_2_1.html", page="7.2 Chấm công tự động thực tế",danhsach=rows)
    
@app.route("/muc8_1", methods=["GET","POST"])
@login_required
@roles_required('user','developer','nhansu','cong','sa','luong','tuyendung')
def ykiemnhamay():

    return render_template("8_1.html", page="8.1 Ý kiến đóng góp nhà máy")
    
@app.route("/muc8_2", methods=["GET","POST"])
@login_required
@roles_required('user','developer','nhansu','cong','sa','luong','tuyendung')
def ykienphanmem():
    
    danhsach = laydanhsachykienphanmem()
    return render_template("8_2.html",
                            page="8.2 Ý kiến đóng góp phần mềm",
                            danhsach=danhsach)

@app.route("/muc8_3", methods=["GET","POST"])
@login_required
@roles_required('user','developer','nhansu','cong','sa','luong','tuyendung')
def ykienkhieunai():
    
    return render_template("8_3.html", page="8.3 Ý kiến khiếu nại")
    
@app.route("/muc9_1", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def xulykiluat():
    
    if request.method == "GET":
        danhsach = laydanhsachkyluat()
        return render_template("9_1.html", page="9.1 Xử lý kỉ luật",danhsach=danhsach)
    else:
        mst = request.form.get("mst")
        hoten = request.form.get("hoten")
        chucvu = request.form.get("chucvu")
        bophan = request.form.get("bophan")
        chuyento = request.form.get("chuyento")
        ngayvao = request.form.get("ngayvao")
        ngayvipham = request.form.get("ngayvipham")
        diadiem = request.form.get("diadiem")
        ngaylapbienban = request.form.get("ngaylapbienban")
        noidung = request.form.get("noidung")
        bienphap = request.form.get("bienphap")
        # print(mst,hoten,chucvu,bophan,chuyento,ngayvao,ngayvipham,diadiem,ngaylapbienban,noidung,bienphap)
        try:
            themdanhsachkyluat(mst,hoten,chucvu,bophan,chuyento,ngayvao,ngayvipham,diadiem,ngaylapbienban,noidung,bienphap)
        except Exception as ex:
            print(ex)
        return redirect("/muc9_1") 
    
@app.route("/muc10_1", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def phongvannghiviec():
        
    return render_template("10_1.html", page="10.1 Tổng hợp phỏng vấn nghỉ việc")

    
@app.route("/muc10_3", methods=["GET","POST"])
@login_required
@roles_required('nhansu','sa','developer')
def inchamduthopdong():
     
    if request.method == "GET":
        return render_template("10_3.html", page="10.3 In chấm dứt hợp đồng")
    elif request.method == "POST":
        mst = request.form.get("mst")
        ngaylamhopdong = request.form.get("ngaylamhd")[-2:]
        thanglamhopdong = request.form.get("ngaylamhd")[5:7]
        namlamhopdong = request.form.get("ngaylamhd")[:4]
        tennhanvien = request.form.get("hoten")
        chucvu = request.form.get("chucvu")
        ngaynghi = request.form.get("ngaynghi")
        try:
            file = inchamduthd(mst,
                    ngaylamhopdong,
                    thanglamhopdong,
                    namlamhopdong,
                    tennhanvien,
                    chucvu,
                    ngaynghi)
            print(f"File: {file}")
            if file:
                return send_file(file, as_attachment=True, download_name="chamduthopdong.xlsx")
            else:
                return redirect("/muc3_3")
        except:
            return redirect("/muc3_3")  


@app.route("/thaydoiphanquyen", methods=["POST"])
def thaydoiphanquyen():
    if request.method == "POST":
        userid= request.form["id"]
        newrole = request.form["newrole"]
        user = Users.query.filter_by(id=userid).first()
        if user:
            user.role = newrole
            db.session.commit()
        return redirect("/admin")
    
@app.route("/update_diemdanhbu", methods=["POST"])
def update_diemdanhbu():
    if request.method == "POST":
        mst = request.form["mst"]
        bophan = request.form["bophan"]
        ngayduyet = request.form["ngayduyet"]
        mstduyet = request.form["mstduyet"]
        loaidiemdanh = request.form["loaidiemdanh"]
        if laycacbophanduocduyet(mstduyet,bophan):
            capnhat_diemdanhbu(mst,ngayduyet,loaidiemdanh)
        return redirect("/muc7_1_4")
    
@app.route("/update_xinnghiphep", methods=["POST"])
def update_xinnghiphep():
    if request.method == "POST":
        mst = request.form["mst"]
        bophan = request.form["bophan"]
        ngayduyet = request.form["ngaynghi"]
        mstduyet = request.form["mstduyet"]
        print(mstduyet,bophan,ngayduyet,mst)
        if laycacbophanduocduyet(mstduyet,bophan):
            capnhat_xinnghiphep(mst,ngayduyet)
        return redirect("/muc7_1_5")
    
@app.route("/taimautangcanhom", methods=["POST"])
def taimautangcanhom():
    if request.method == "POST":
        phongban = request.form["phongban"]
        users = laydanhsachusertheophongban(phongban)
        result = []
        ngaydangky = datetime.now().date()
        for user in users:
            result.append({
                "MST": user["MST"],
                "Họ tên": user["Họ tên"],
                "Chức vụ": user["Chức vụ"],
                "Chuyền tổ": user["Line"], 
                "Phòng ban": user["Department"],
                "Ngày đăng ký": ngaydangky,
                "Giờ tăng ca": "18:30",
                "Giờ tăng ca thực tế":"20:30"
            })
        df = pd.DataFrame(result)
        df.to_excel("tangca.xlsx", index=False)
        
        return send_file("tangca.xlsx", as_attachment=True)  
    
@app.route("/capnhattrangthaiungvien", methods=["POST"])
def capnhattrangthaiungvien():
    print(request.args)
    sdt = request.args.get("sdt")
    trangthai = request.args.get("trangthaimoi")
    if sdt and trangthai:
        capnhattrangthai(sdt, trangthai)
        return {"status": "success"}, 200
    return {"status": "fail"}, 400

@app.route("/laythongtincccd", methods=["POST"])
def laythongtincccd():
    
    conn = pyodbc.connect(used_db)
    cursor = conn.cursor()

    if request.method == "POST":
        cccd = request.args.get("cccd")  # lấy giá trị cccd từ form data
        if cccd:
            employee = cursor.execute("SELECT * FROM HR.dbo.Dang_ky_thong_tin WHERE CCCD = ?", cccd).fetchone()
            conn.close()
            if employee:
                tamtru = employee[10]
                data_tamtru = tamtru.split(",")
                
                # Chuyển đổi employee thành dict và trả về dạng JSON
                employee_dict = {
                    "Nhà máy": employee[0],
                    "Vị trí ứng tuyển": employee[1],
                    "Số điện thoại": employee[3],
                    "Số CCCD": employee[4],
                    "Dân tộc": employee[5],
                    "Quốc tịch": employee[7],
                    "Tôn giáo": employee[6],
                    "Trình độ học vấn": employee[8],
                    "Nơi sinh" : employee[9],
                    "Tạm trú" : tamtru,
                    "Phường/Xã": data_tamtru[1] if len(data_tamtru) > 1 else "",
                    "Quận/huyện": data_tamtru[2] if len(data_tamtru) > 2 else "",
                    "Tỉnh/Thành phố": data_tamtru[3] if len(data_tamtru) > 3 else "",
                    "Số BHXH": employee[11],
                    "Mã số thuế": employee[12],
                    "Ngân hàng": employee[13],
                    "Số tài khoản": employee[14],
                    "Con nhỏ": employee[21],
                    "Tên con 1": employee[22],
                    "Ngày sinh con 1": employee[23],
                    "Tên con 2": employee[24],
                    "Ngày sinh con 2": employee[24],
                    "Tên con 3": employee[25],
                    "Ngày sinh con 3": employee[25],
                    "Tên con 4": employee[26],
                    "Ngày sinh con 4": employee[26],
                    "Tên con 5": employee[27],
                    "Ngày sinh con 5": employee[27],
                }
                return jsonify(employee_dict)
            else:
                return jsonify({"error": "Employee not found"}), 404
        else:
            return jsonify({"error": "CCCD is required"}), 400

@app.route("/kiemtrathongtinnld", methods=["POST"])
def kiemtrathongtinnld():

    if request.method == "POST":
        mst = request.args.get("masothe")
        if mst:
            users = laydanhsachtheomst(mst)
            if users:
                return jsonify(users[0]), 200
            else:
                return jsonify({"error": "User not found"}), 404
        else:
            return jsonify({"error": "MST is required"}), 400

@app.route("/thaydoithongtin_hopdong", methods=["POST"])
def thaydoithongtin_hopdong():
    
    if request.method == "POST":
        loaihopdong = request.args.get("loaihopdong")
        mst = request.args.get("mst")
        ngaylamhopdong = request.args.get("ngaybatdau")
        thanglamhopdong = request.args.get("thangbatdau")
        namlamhopdong = request.args.get("nambatdau")
        ngayketthuchopdong = request.args.get("ngayketthuc")
        thangketthuchopdong = request.args.get("thangketthuc")
        namketthuchopdong = request.args.get("namketthuc")
        tennhanvien = request.args.get("hoten")
        ngaysinh = request.args.get("ngaysinh")
        gioitinh = request.args.get("gioitinh")
        thuongtru = request.args.get("thuongtru")
        cccd = request.args.get("cccd")
        ngaycapcccd = request.args.get("ngaycapcccd")
        mucluong = request.args.get("luongcoban")
        chucvu = request.args.get("chucvu")
        try:
            file = thaydoithongtinhopdong(loaihopdong, mst,ngaylamhopdong,thanglamhopdong,namlamhopdong,ngayketthuchopdong,thangketthuchopdong,namketthuchopdong,tennhanvien,ngaysinh,gioitinh,thuongtru,cccd,ngaycapcccd,mucluong,chucvu)
            if file:
                print(file)
                return send_file(file, as_attachment=True)
            else:
                return jsonify({"error": "Loi in hop dong"}), 500
        except Exception as e:
            return jsonify({"error": str(e)}), 500

@app.route("/dangkitangcacanhan", methods=["POST"])  
def dangkitangcacanhan():
    mst = request.form.get("mst")
    giotangca = request.form.get("giotangca")
    giotangcathucte = request.form.get("giotangcathucte")
    ngaytangca = request.form.get("ngaytangca")
    thongtin = laydanhsachtheomst(mst)[0]
    insert_tangca(current_user.macongty,mst,thongtin['Họ tên'],thongtin['Chức vụ'],thongtin['Line'],thongtin['Department'],ngaytangca,giotangca, giotangcathucte)
    return redirect(f"/muc7_1_6?ngay={ngaytangca}")
    
    
@app.route("/dangkitangcanhom", methods=["POST"])   
def dangkitangcanhom():
    
    if request.method == "POST":
        try:
            if 'file' not in request.files:
                return redirect("/muc7_1_6")
            file = request.files['file']
            if file.filename == '':
                return redirect("/muc7_1_6")
            if file:
                filename = secure_filename(file.filename)
                ngaylam = datetime.now().strftime("%d%m%Y_%H%M%S")
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"tangca_{current_user.phongban}_{ngaylam}.xlsx")
                file.save(filepath)
                data = pd.read_excel(filepath).to_dict(orient="records")
                for row in data:
                    try:
                        insert_tangca(current_user.macongty,row["MST"],row["Họ tên"],row["Chức vụ"],row["Chuyền tổ"],row["Phòng ban"],row["Ngày đăng ký"],row["Giờ tăng ca"], row["Giờ tăng ca thực tế"])
                    except Exception as e:
                        print(e)                
                return redirect("/muc7_1_6")
        except Exception as e:
            print(e)
            return redirect("/muc7_1_6")

@app.route("/export_dstc", methods=["POST"])
def export_dstc():
    mst = request.form.get("mst")
    phongban = request.form.get("phongban")
    ngay = request.form.get("ngay")
    danhsach = laydanhsachtangca(mst,phongban,ngay)
    result = []
    for row in danhsach:
        result.append(
            {
                'Nhà máy': row[0],
                'MST': row[1],
                'Họ tên': row[2],
                'Chức danh': row[3],
                'Chuyền': row[4],
                'Phòng ban': row[5],
                'Ngày đăng ký': row[6],
                'Giờ tăng ca': row[7],
                'Giờ tăng ca thực tế': row[8],
            }
        )
    df = pd.DataFrame(result)
    thoigian = datetime.now().strftime("%d%m%Y_%H%M%S")
    df.to_excel(f"danhsachtangca_{thoigian}.xlsx", index=False)
    
    return send_file(f"danhsachtangca_{thoigian}.xlsx", as_attachment=True)
    
       
@app.route("/export_dsnv", methods=["POST"])
def export_dsnv():
    
    mst = request.form.get("mst")
    hoten = request.form.get("hoten")
    sdt = request.form.get("sdt")
    cccd = request.form.get("cccd")
    gioitinh = request.form.get("gioitinh")
    ngayvao = request.form.get("ngayvao")
    ngaynghi = request.form.get("ngaynghi")
    ngaykyhd = request.form.get("ngaykyhd")
    ngayhethanhd = request.form.get("ngayhethanhd")
    phongban = request.form.get("phongban")
    trangthai = request.form.get("trangthai")
    hccategory = request.form.get("hccategory")
    users = laydanhsachuser(mst, hoten, sdt, cccd, gioitinh, ngayvao, ngaynghi, ngaykyhd, ngayhethanhd, phongban, trangthai, hccategory)    
    df = pd.DataFrame(users)
    df.to_excel("danhsach.xlsx", index=False)
    
    return send_file("danhsach.xlsx", as_attachment=True)  

@app.route("/export_dslt", methods=["POST"])
def export_dslt():
    
    chuyen = request.form.get("chuyen")
    bophan = request.form.get("bophan")
    ngay = request.form.get("ngay")
    rows = laydanhsachloithe(chuyen, bophan, ngay)
    
    df = pd.DataFrame(rows)
    df.to_excel("loithe.xlsx", index=False)
    
    return send_file("loithe.xlsx", as_attachment=True)  

@app.route("/export_dsddb", methods=["POST"])
def export_dsddb():
    
    mst = request.form.get("mst")
    chuyen = request.form.get("chuyen")
    bophan = request.form.get("bophan")
    rows = laydanhsachdiemdanhbu(mst,chuyen, bophan)
    
    df = pd.DataFrame(rows)
    df.to_excel("diemdanhbu.xlsx", index=False)
    
    return send_file("diemdanhbu.xlsx", as_attachment=True)  

@app.route("/export_dsxnp", methods=["POST"])
def export_dsxnp():
    
    mst = request.form.get("mst")
    chuyen = request.form.get("chuyen")
    bophan = request.form.get("bophan")
    rows = laydanhsachxinnghiphep(mst, chuyen, bophan)
    
    df = pd.DataFrame(rows)
    df.to_excel("xinnghiphep.xlsx", index=False)
    
    return send_file("xinnghiphep.xlsx", as_attachment=True)  

@app.route("/export_dsdktt", methods=["POST"])
def export_dsdktt():
    
    sdt = request.form.get("sdt")
    cccd = request.form.get("cccd")
    rows = laydanhsachdangkytuyendung(sdt, cccd)   
    df = pd.DataFrame(rows)
    df.to_excel("tuyendung.xlsx", index=False)
    
    return send_file("tuyendung.xlsx", as_attachment=True)  
      
@app.route("/check_hcname", methods=["POST"])
def check_hcname():
    jobtitle = request.args.get("vitri")
    line = request.args.get("line")
    hcname = layhcname(jobtitle,line)
    if not hcname:
        return jsonify({
            "Line": "",
            "Detail_job_title_VN": "",
            "Detail_job_title_EN": "",
            "Employee_type": "",
            "Position_code": "",
            "Position_code_description": "",
            "Grade_code": "",
            "HC_category": "",
            "Factory": "",
            "Department": "",
            "Section_code": "",
            "Section_description": ""
        })
    return jsonify({
        "Line": hcname[0],
        "Detail_job_title_VN": hcname[1],
        "Detail_job_title_EN": hcname[2],
        "Employee_type": hcname[3],
        "Position_code": hcname[4],
        "Position_code_description": hcname[5],
        "Grade_code": hcname[6],
        "HC_category": hcname[7],
        "Factory": hcname[8],
        "Department": hcname[9],
        "Section_code": hcname[10],
        "Section_description": hcname[11]        
    })

@app.route("/check_line_from_detailjob", methods=["POST"])
def check_line_from_detailjob():
    vitri = request.args.get("vitrimoi")
    cacline = laydanhsachlinetheovitri(vitri)
    return jsonify(cacline)

@app.route("/xoanhanviencu", methods=["GET"])
def xoanhanviencu():
    mst = request.args.get("mst")
    try:
        print(xoanhanvien(mst))
        return redirect(url_for('timdanhsachnhanvien', mst=mst))
    except Exception as e:
        print(e)
        return redirect(url_for('timdanhsachnhanvien', mst=mst))

@app.route("/doicacanhan", methods=["POST"])
def doicacanhan():
    mst = request.form.get("mst")
    cacu = request.form.get("cacu")
    camoi = request.form.get("camoi")
    themdoicamoi(mst,cacu,camoi)
    return redirect("/muc7_1_1")

@app.route("/doicanhom", methods=["POST"])
def doicanhom():
    phongban = request.form.get("phongban")
    cacu = request.form.get("cacu")
    camoi = request.form.get("camoi")
    danhsach = laydanhsachusertheophongban(phongban)
    for user in danhsach:
        themdoicamoi(user['MST'],cacu,camoi)
    return redirect("/muc7_1_1")

@app.route("/laycatheomst", methods=["POST"])
def laycatheomst():
    mst = request.args.get("mst")
    ca = laycahientai(mst)
    return jsonify({
        "Ca": ca
    })
    

if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=81
    )