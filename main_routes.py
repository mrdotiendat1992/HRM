# -*- encoding: utf-8 -*-

import openpyxl.styles
from app import *

##################################
#          MAIN ROUTES           #
##################################

@app.before_request
def run_before_every_request():
    try:
        if current_user.is_authenticated:
            f12 = trang_thai_function_12()
            g.notice={"f12":f12, "db":url_database_pyodbc ,"Tổng":0}
            
            # Kiểm tra xem có phải quản lý không
            laquanly = la_quanly(current_user.macongty,current_user.masothe)
            if laquanly:
                quanly_soluong_diemdanhbu = lay_soluong_diemdanhbu_quanly_canduyet(current_user.macongty,current_user.masothe)
                quanly_soluong_xinnghiphep = lay_soluong_xinnghiphep_quanly_canduyet(current_user.macongty,current_user.masothe)
                quanly_soluong_xinnghikhongluong = lay_soluong_xinnghikhongluong_quanly_canduyet(current_user.macongty,current_user.masothe)
                quanly_soluong_xinnghikhac = lay_soluong_xinnghikhac_quanly_canduyet(current_user.macongty,current_user.masothe)
                g.notice["Quản lý"]={"Điểm danh bù":quanly_soluong_diemdanhbu,
                    "Xin nghỉ phép": quanly_soluong_xinnghiphep,
                    "Xin nghỉ không lương": quanly_soluong_xinnghikhongluong,
                    "Xin nghỉ khác": quanly_soluong_xinnghikhac,
                    "Số thông báo": quanly_soluong_diemdanhbu + quanly_soluong_xinnghiphep + quanly_soluong_xinnghikhongluong + quanly_soluong_xinnghikhac
                    }
                g.notice["Tổng"] = g.notice["Tổng"] + quanly_soluong_diemdanhbu + quanly_soluong_xinnghiphep + quanly_soluong_xinnghikhongluong + quanly_soluong_xinnghikhac
            else:
                g.notice["Quản lý"]={}
            lathuky= la_thuky(current_user.macongty,current_user.masothe)
            if lathuky:
                cac_chuyen_thuky_quanly = lay_danhsach_chuyen_thuky_quanly(current_user.macongty,current_user.masothe)
                # print(cac_chuyen_thuky_quanly)
                thuky_soluong_loithe = lay_soluong_loithe_thuky_canxuly(current_user.macongty,current_user.masothe)
                thuky_soluong_diemdanhbu = lay_soluong_diemdanhbu_thuky_cankiemtra(current_user.macongty,current_user.masothe)
                thuky_soluong_xinnghiphep = lay_soluong_xinnghiphep_thuky_cankiemtra(current_user.macongty,current_user.masothe)
                thuky_soluong_xinnghikhongluong = lay_soluong_xinnghikhongluong_thuky_cankiemtra(current_user.macongty,current_user.masothe)
                thuky_soluong_xinnghikhac = lay_soluong_xinnghikhac_thuky_cankiemtra(current_user.macongty,current_user.masothe)
                
                g.notice["Thư ký"]={"Danh sách lỗi thẻ":thuky_soluong_loithe,
                                    "Điểm danh bù":thuky_soluong_diemdanhbu,
                                    "Xin nghỉ phép": thuky_soluong_xinnghiphep,
                                    "Xin nghỉ không lương": thuky_soluong_xinnghikhongluong,
                                    "Xin nghỉ khác": thuky_soluong_xinnghikhac,
                                    "Line":cac_chuyen_thuky_quanly[0] if len(cac_chuyen_thuky_quanly)==1 else "",
                                    "Số thông báo":thuky_soluong_loithe + thuky_soluong_diemdanhbu + thuky_soluong_xinnghiphep + thuky_soluong_xinnghikhongluong}
                g.notice["Tổng"] = g.notice["Tổng"] + thuky_soluong_loithe + thuky_soluong_diemdanhbu + thuky_soluong_xinnghiphep + thuky_soluong_xinnghikhongluong + thuky_soluong_xinnghikhac
            else:
                g.notice["Thư ký"]={}

            so_don_diemdanhbu_chuakiemtra = lay_soluong_diemdanhbu_chuakiemtra(current_user.macongty,current_user.masothe)
            so_don_diemdanhbu_dakiemtra = lay_soluong_diemdanhbu_dakiemtra(current_user.macongty,current_user.masothe)
            so_don_diemdanhbu_dapheduyet = lay_soluong_diemdanhbu_dapheduyet(current_user.macongty,current_user.masothe)
            so_don_diemdanhbu_bituchoi = lay_soluong_diemdanhbu_bituchoi(current_user.macongty,current_user.masothe)        
            so_don_diemdanhbu = sum([so_don_diemdanhbu_chuakiemtra,so_don_diemdanhbu_dakiemtra,so_don_diemdanhbu_dapheduyet,so_don_diemdanhbu_bituchoi])

            so_don_xinnghiphep_chuakiemtra = lay_soluong_xinnghiphep_chuakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghiphep_dakiemtra = lay_soluong_xinnghiphep_dakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghiphep_dapheduyet = lay_soluong_xinnghiphep_dapheduyet(current_user.macongty,current_user.masothe)
            so_don_xinnghiphep_bituchoi = lay_soluong_xinnghiphep_bituchoi(current_user.macongty,current_user.masothe)        
            so_don_xinnghiphep = sum([so_don_xinnghiphep_chuakiemtra,so_don_xinnghiphep_dakiemtra,so_don_xinnghiphep_dapheduyet,so_don_xinnghiphep_bituchoi])
            
            so_don_xinnghikhongluong_chuakiemtra = lay_soluong_xinnghikhongluong_chuakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghikhongluong_dakiemtra = lay_soluong_xinnghikhongluong_dakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghikhongluong_dapheduyet = lay_soluong_xinnghikhongluong_dapheduyet(current_user.macongty,current_user.masothe)
            so_don_xinnghikhongluong_bituchoi = lay_soluong_xinnghikhongluong_bituchoi(current_user.macongty,current_user.masothe)        
            so_don_xinnghikhongluong = sum([so_don_xinnghikhongluong_chuakiemtra,so_don_xinnghikhongluong_dakiemtra,so_don_xinnghikhongluong_dapheduyet,so_don_xinnghikhongluong_bituchoi])
            
            so_don_xinnghikhac_chuakiemtra = lay_soluong_xinnghikhac_chuakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghikhac_dakiemtra = lay_soluong_xinnghikhac_dakiemtra(current_user.macongty,current_user.masothe)
            so_don_xinnghikhac_dapheduyet = lay_soluong_xinnghikhac_dapheduyet(current_user.macongty,current_user.masothe)
            so_don_xinnghikhac_bituchoi = lay_soluong_xinnghikhac_bituchoi(current_user.macongty,current_user.masothe)        
            so_don_xinnghikhac = so_don_xinnghikhac_chuakiemtra + so_don_xinnghikhac_dakiemtra + so_don_xinnghikhac_dapheduyet + so_don_xinnghikhac_bituchoi
            
            so_don = so_don_diemdanhbu + so_don_xinnghiphep + so_don_xinnghikhongluong + so_don_xinnghikhac
            so_lan_loi_cham_cong = lay_soluong_loichamcong(current_user.macongty,current_user.masothe)                
            
            g.notice["personal"]={
                "Điểm danh bù":{
                                "Chưa kiểm tra":so_don_diemdanhbu_chuakiemtra,
                                "Đã kiểm tra": so_don_diemdanhbu_dakiemtra,
                                "Đã phê duyệt": so_don_diemdanhbu_dapheduyet,
                                "Bị từ chối": so_don_diemdanhbu_bituchoi,
                                "Tổng": so_don_diemdanhbu
                            },
                "Xin nghỉ phép":{
                                "Chưa kiểm tra":so_don_xinnghiphep_chuakiemtra,
                                "Đã kiểm tra": so_don_xinnghiphep_dakiemtra,
                                "Đã phê duyệt": so_don_xinnghiphep_dapheduyet,
                                "Tổng": so_don_xinnghiphep,
                                "Bị từ chối": so_don_xinnghiphep_bituchoi,
                            },
                "Xin nghỉ không lương":{
                                "Chưa kiểm tra":so_don_xinnghikhongluong_chuakiemtra,
                                "Đã kiểm tra": so_don_xinnghikhongluong_dakiemtra,
                                "Đã phê duyệt": so_don_xinnghikhongluong_dapheduyet,
                                "Tổng": so_don_xinnghikhongluong,
                                "Bị từ chối": so_don_xinnghikhongluong_bituchoi,
                            },
                "Xin nghỉ khác":{
                                "Chưa kiểm tra":so_don_xinnghikhac_chuakiemtra,
                                "Đã kiểm tra": so_don_xinnghikhac_dakiemtra,
                                "Đã phê duyệt": so_don_xinnghikhac_dapheduyet,
                                "Tổng": so_don_xinnghikhac,
                                "Bị từ chối": so_don_xinnghikhac_bituchoi,
                            },
                "Tổng":so_don,
                "Lỗi chấm công": so_lan_loi_cham_cong
                                }

            if current_user.phanquyen=='gd':
                soluong_yeucautuyendung_chopheduyet = lay_soluong_yeucautuyendung_chopheduyet(current_user.macongty)
                if soluong_yeucautuyendung_chopheduyet > 0:
                    g.notice["Tuyển dụng chờ phê duyệt"] = soluong_yeucautuyendung_chopheduyet
                    g.notice["Tổng"] += soluong_yeucautuyendung_chopheduyet
                else:
                    g.notice["Tuyển dụng chờ phê duyệt"] = 0

            if current_user.phanquyen=='tbp' or current_user.phanquyen=='sa':
                soluong_yeucautuyendung_dapheduyet = lay_soluong_yeucautuyendung_dapheduyet(current_user.macongty,current_user.masothe)
                if soluong_yeucautuyendung_dapheduyet > 0:
                    g.notice["Tuyển dụng được duyệt"] = soluong_yeucautuyendung_dapheduyet
                    g.notice["Tổng"] += soluong_yeucautuyendung_dapheduyet
                else:
                    g.notice["Tuyển dụng được duyệt"] = 0
                soluong_yeucautuyendung_bituchoi = lay_soluong_yeucautuyendung_bituchoi(current_user.macongty,current_user.masothe)
                if soluong_yeucautuyendung_bituchoi > 0:
                    g.notice["Tuyển dụng bị từ chối"] = soluong_yeucautuyendung_bituchoi   
                    g.notice["Tổng"] += soluong_yeucautuyendung_bituchoi
                else:
                    g.notice["Tuyển dụng bị từ chối"] = 0
        # print(g.notice)
    except Exception as e:  
        flash(f"Lỗi cập nhật thông tin lỗi chấm công: {e}")
        f12 = trang_thai_function_12()    
        g.notice={"f12":f12,"db":url_database_pyodbc }
    
@app.context_processor
def inject_notice():
    return dict(notice=getattr(g, 'notice', {}),personal = getattr(g, 'personal', {}))

@app.route('/unauthorized')
def unauthorized():
    return render_template_string("<h1>Bạn không thể vào mục này, vui lòng chọn mục khác!!!</h1><h3>Ấn vào <a href='/'>đây</a> để quay lại trang chủ</h3>")

@app.errorhandler(404)
def page_not_found(e):
    return render_template('blank.html'), 404
 
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        try:
            macongty = request.form['macongty']
            masothe = request.form['masothe']
            matkhau = request.form['matkhau']
            user = Nhanvien.query.filter_by(masothe=masothe, macongty=macongty).first()    
            if user and user.matkhau == matkhau:
                if login_user(user):    
                    app.logger.info(f"Nguoi dung {masothe} o {macongty} vua  dang nhap thanh cong !!!")
                    return redirect(url_for('home'))
            return redirect(url_for("login"))
        except Exception as e:
            app.logger.error(f'Nguoi dung {masothe} o {macongty} dang nhap that bai: {e} !!!')
            return redirect(url_for("login"))
    return render_template("login.html")

@app.route("/logout", methods=["POST"])
def logout():
    try:
        app.logger.info(f"Nguoi dung {current_user.masothe} o {current_user.macongty} vua  dang xuat !!!")
        logout_user()
    except Exception as e:
        app.logger.error(f'Không thế đăng xuất {e} !!!')
    return redirect("/")

@app.route("/doimatkhau", methods=['POST'])
def doimatkhau():
    macongty = request.form.get("macongty")
    masothe = request.form.get("masothe_doi")
    matkhaumoi = request.form.get("matkhaumoi")
    try:
        if doimatkhautaikhoan(macongty,masothe,matkhaumoi):
            flash("Đổi mật khẩu thành công")
    except Exception as e:
        app.logger.error(f"{masothe} o {macongty} doi mat khau thanh {matkhaumoi} thanh cong !!!")
    return redirect(url_for("home"))

@app.route("/home", methods=['GET','POST'])
@login_required
def index():
    return redirect("/")

@app.route("/", methods=['GET','POST'])
@login_required
def home():
    if request.method == "GET":
        mst = request.args.get("Mã số thẻ")
        hoten = request.args.get("Họ tên")
        sdt = request.args.get("Số điện thoại")
        cccd = request.args.get("Căn cước công dân")
        gioitinh = request.args.get("Giới tính")
        vaotungay = request.args.get("Vào từ ngày")
        vaodenngay = request.args.get("Vào đến ngày")
        nghitungay = request.args.get("Nghỉ từ ngày")
        nghidenngay = request.args.get("Nghỉ đến ngày")
        phongban = request.args.get("Phòng ban")
        chucvu = request.args.get("Chức danh")
        trangthai = request.args.get("Trạng thái")
        hccategory = request.args.get("HC Category")
        ghichu = request.args.get("Ghi chú")
        chuyen = request.args.get("Chuyền")
        users = laydanhsachuser(mst, hoten, sdt, cccd, gioitinh, vaotungay, vaodenngay, nghitungay, nghidenngay, phongban, trangthai, hccategory, chucvu, ghichu, chuyen)   
        count = len(users)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(users)
        start = (page - 1) * per_page
        end = start + per_page
        paginated_users = users[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        songuoi_danglamviec = lay_soluong_danglamviec()
        songuoi_dangnghithaisan = lay_soluong_dangnghithaisan()
        flash(f"Xin chào {current_user.hoten} !!!")
        return render_template("home.html", users=paginated_users,
                            page="Trang chủ", pagination=pagination,count=count,
                            songuoi_danglamviec=songuoi_danglamviec,
                            songuoi_dangnghithaisan=songuoi_dangnghithaisan)
    else:
        try:
            mst = request.form.get("Mã số thẻ")
            hoten = request.form.get("Họ tên")
            sdt = request.form.get("Số điện thoại")
            cccd = request.form.get("Căn cước công dân")
            gioitinh = request.form.get("Giới tính")
            vaotungay = request.form.get("Vào từ ngày")
            vaodenngay = request.form.get("Vào đến ngày")
            nghitungay = request.form.get("Nghỉ từ ngày")
            nghidenngay = request.form.get("Nghỉ đến ngày")
            phongban = request.form.get("Phòng ban")
            chucvu = request.form.get("Chức danh")
            trangthai = request.form.get("Trạng thái")
            hccategory = request.form.get("Headcount Category")
            ghichu = request.form.get("Ghi chú")
            chuyen = request.form.get("Chuyền")
            users = laydanhsachuser(mst, hoten, sdt, cccd, gioitinh, vaotungay, vaodenngay, nghitungay, nghidenngay, phongban, trangthai, hccategory, chucvu, ghichu, chuyen)      
            
            # Chuyển thông tin ngày về định dạng YYYY-MM-DD
            for user in users:
                user["Ngày sinh"] = datetime.strptime(user["Ngày sinh"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày sinh"]!="" else ""
                user["Ngày cấp CCCD"] = datetime.strptime(user["Ngày cấp CCCD"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày cấp CCCD"]!="" else ""
                user["Ngày ký HĐ"] = datetime.strptime(user["Ngày ký HĐ"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày ký HĐ"]!="" else ""
                user["Ngày vào"] = datetime.strptime(user["Ngày vào"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày vào"]!="" else ""
                user["Ngày nghỉ"] = datetime.strptime(user["Ngày nghỉ"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày nghỉ"]!="" else ""
                user["Ngày hết hạn"] = datetime.strptime(user["Ngày hết hạn"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày hết hạn"]!="" else ""
                user["Ngày vào nối thâm niên"] = datetime.strptime(user["Ngày vào nối thâm niên"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày vào nối thâm niên"]!="" else ""
                user["Ngày kí HĐ Thử việc"] = datetime.strptime(user["Ngày kí HĐ Thử việc"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày kí HĐ Thử việc"]!="" else ""
                user["Ngày hết hạn HĐ Thử việc"] = datetime.strptime(user["Ngày hết hạn HĐ Thử việc"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày hết hạn HĐ Thử việc"]!="" else ""
                user["Ngày hết hạn HĐ xác định thời hạn lần 1"] = datetime.strptime(user["Ngày hết hạn HĐ xác định thời hạn lần 1"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày hết hạn HĐ xác định thời hạn lần 1"]!="" else ""
                user["Ngày kí HĐ xác định thời hạn lần 1"] = datetime.strptime(user["Ngày kí HĐ xác định thời hạn lần 1"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày kí HĐ xác định thời hạn lần 1"]!="" else ""
                user["Ngày kí HĐ không thời hạn"] = datetime.strptime(user["Ngày kí HĐ không thời hạn"],"%d/%m/%Y").strftime("%Y-%m-%d") if user["Ngày kí HĐ không thời hạn"]!="" else ""
                

            df = pd.DataFrame(users)

            df["Ngày sinh"] = to_datetime(df['Ngày sinh'])
            df["Ngày cấp CCCD"] = to_datetime(df['Ngày cấp CCCD'])
            df["Ngày ký HĐ"] = to_datetime(df['Ngày ký HĐ'])
            df["Ngày vào"] = to_datetime(df['Ngày vào'])
            df["Ngày nghỉ"] = to_datetime(df['Ngày nghỉ'])
            df["Ngày hết hạn"] = to_datetime(df['Ngày hết hạn'])
            df["Ngày vào nối thâm niên"] = to_datetime(df['Ngày vào nối thâm niên'])
            df["Ngày sinh con 1"] = to_datetime(df['Ngày sinh con 1'])
            df["Ngày sinh con 2"] = to_datetime(df['Ngày sinh con 2'])
            df["Ngày sinh con 3"] = to_datetime(df['Ngày sinh con 3'])
            df["Ngày sinh con 4"] = to_datetime(df['Ngày sinh con 4'])
            df["Ngày sinh con 5"] = to_datetime(df['Ngày sinh con 5'])
            df["Ngày kí HĐ Thử việc"] = to_datetime(df['Ngày kí HĐ Thử việc'])
            df["Ngày hết hạn HĐ Thử việc"] = to_datetime(df['Ngày hết hạn HĐ Thử việc'])
            df["Ngày kí HĐ xác định thời hạn lần 1"] = to_datetime(df['Ngày kí HĐ xác định thời hạn lần 1'])
            df["Ngày hết hạn HĐ xác định thời hạn lần 1"] = to_datetime(df['Ngày hết hạn HĐ xác định thời hạn lần 1'])
            df["Ngày kí HĐ không thời hạn"] = to_datetime(df['Ngày kí HĐ không thời hạn'])
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False)

            # Adjust column width and format the header row
            output.seek(0)
            workbook = openpyxl.load_workbook(output)
            sheet = workbook.active

            # Style the header row
            header_fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font

            # Create a date format for short date
            date_format = NamedStyle(name="short_date", number_format="DD/MM/YYYY")
            if "short_date" not in workbook.named_styles:
                workbook.add_named_style(date_format)
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                        if cell.column_letter in ['E','H','AB','AD','AF','AF','AJ','AO','AP','BG','BH','BJ','BL','BM','BM','BO','BP','BQ','BR'] and cell.value is not None:
                            cell.number_format = 'DD/MM/YYYY'
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

            # Save the modified workbook to the output BytesIO object
            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            
            # Generate the timestamp for the filename
            time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
            
            # Return the file to the client
            response = make_response(output.read())
            response.headers['Content-Disposition'] = f'attachment; filename=danhsach_nhanvien_{time_stamp}.xlsx'
            response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            return response
        except Exception as e:
            flash(f"Lỗi kết xuất danh sách nhân viên")

@app.route("/muc2_1", methods=["GET","POST"])
@login_required
@roles_required('hr','tnc','sa','gd','td','tbp')
def danhsachdangkytuyendung():
    if request.method == "GET":
        sdt = request.args.get("sdt")
        cccd = request.args.get("cccd")
        ngaygui = request.args.get("ngaygui")
        rows = laydanhsachdangkytuyendung(sdt,cccd,ngaygui)
        count=len(rows)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')

        return render_template("2_1.html", 
                            page="2.1 Danh sách ứng viên",
                            danhsach=paginated_rows, 
                            pagination=pagination,
                            count=count)
        
    if request.method == "POST":
        id = request.form.get("id")
        sdt = request.form.get("sdt")
        vitrituyendung = request.form.get("vitrituyendung")
        hocvan = request.form.get("hocvan")
        diachi = request.form.get("diachi")
        ngayhendilam = request.form.get("ngayhendilam")
        hieusuat = request.form.get("hieusuat")
        loaimay = request.form.get("loaimay")
        cccd = request.form.get("cccd")
        dantoc = request.form.get("dantoc")
        connho = request.form.get("connho")
        tencon1 = request.form.get("tenconnho1")
        ngaysinhcon1 = request.form.get("ngaysinhcon1")
        tencon2 = request.form.get("tenconnho2")
        ngaysinhcon2 = request.form.get("ngaysinhcon2")
        tencon3 = request.form.get("tenconnho3")
        ngaysinhcon3 = request.form.get("ngaysinhcon3")
        tencon4 = request.form.get("tenconnho4")
        ngaysinhcon4 = request.form.get("ngaysinhcon4")
        tencon5 = request.form.get("tenconnho5")
        ngaysinhcon5 = request.form.get("ngaysinhcon5")
        nguoithan = request.form.get("nguoithan")
        sdtnguoithan = request.form.get("sdtnguoithan")
        ngayhendilam = request.form.get("ngayhendilam")
        luuhoso = request.form.get("luuhoso")
        ghichu = request.form.get("ghichu")
        cccd = request.form.get("cccd")
        ketqua = capnhatthongtinungvien(id,
                               sdt,
                               ngayhendilam,
                               hieusuat,
                               loaimay,
                               vitrituyendung,
                               hocvan,
                               diachi,
                               dantoc,
                               connho,
                               tencon1,
                               ngaysinhcon1,
                               tencon2,
                               ngaysinhcon2,
                               tencon3,
                               ngaysinhcon3,
                               tencon4,
                               ngaysinhcon4,
                               tencon5,
                               ngaysinhcon5,
                               nguoithan,
                               sdtnguoithan,
                               luuhoso,
                               ghichu,
                               cccd
                               )
        if ketqua["ketqua"]:
            flash("Cập nhật thông tin ứng viên thành công !!!")
        else:
            flash(f"Cập nhật thông tin ứng viên thất bại, lí do: {ketqua["lido"]}, query: {ketqua["query"]}")
        return redirect(f"muc2_1?sdt={sdt}")

@app.route("/muc2_2", methods=["GET","POST"])
@login_required
@roles_required('tbp','gd','sa','td')
def dangkytuyendung():
    if request.method == "GET":
        phongban = request.args.get("phongban")
        trangthaiyeucau = request.args.get("trangthaiyeucau")
        trangthaithuchien = request.args.get("trangthaithuchien")
        mst = request.args.get("mst")
        danhsach = laydanhsachyeucautuyendung(current_user.macongty,phongban,trangthaiyeucau,trangthaithuchien,mst)
        cactrangthaithuchien = ["Chưa tuyển","Đã đăng tuyển","Chờ phỏng vấn","Đã tuyển"]
        cacvitri = lay_cac_vitri_trong_phong(current_user.phongban)
        return render_template("2_2.html", 
                               page="2.2 Yêu cầu tuyển dụng",
                               danhsach=danhsach,
                               cacvitri = cacvitri,
                               cactrangthaithuchien=cactrangthaithuchien
                               )
    
    elif request.method == "POST":
        try:
            bophan = current_user.phongban
            vitri = request.form.get("vitri")
            vitrien = request.form.get("vitrien")
            capbac = request.form.get("capbac")
            bacluongtu = request.form.get("bacluongtu")
            bacluongden = request.form.get("bacluongden")
            bacluong = f"{bacluongtu.split(",")[0]} => {bacluongden.split(",")[0]}"
            soluong = request.form.get("soluong")
            mota = os.path.join(FOLDER_JD, f"{vitrien}.pdf")
            thoigiandukien = request.form.get("thoigiandukien")
            phanloai = request.form.get("phanloai")
            khoangluong = f"{bacluongtu.split(",")[1]} => {bacluongden.split(",")[1]}"
            if themyeucautuyendungmoi(bophan,vitri,soluong,mota,thoigiandukien,phanloai,khoangluong,capbac,bacluong):
                flash("Thêm yêu cầu tuyển dụng mới thành công !!!")
                if them_thongbao_co_yeucautuyendung(current_user.masothe,current_user.hoten):
                    flash("Thêm thông báo có yêu cầu tuyển dụng mới thành công !!!")
                else:
                    flash("Thêm thông báo có yêu cầu tuyển dụng mới thất bại !!!")
                
            else:
                flash("Thêm yêu cầu tuyển dụng mới thất bại !!!")
        except Exception as e:
            flash(f"Thêm yêu cầu tuyển dụng mới thất bại ({e})!!!")
        return redirect("muc2_2")

@app.route("/muc2_2_1", methods=["GET","POST"])
@login_required
@roles_required('tbp','gd','sa','td')
def tuyendungchitiet():
    if request.method == "GET":
        id_yeucautuyendung = request.args.get("id")
        danhsach = lay_danhsach_ungvien(id_yeucautuyendung)
        phongban = lay_phongban_theo_idyctd(id_yeucautuyendung)
        return render_template("2_2_1.html", 
                               page="2.2.1 Danh sách ứng viên tuyển dụng",
                               danhsach=danhsach,
                               phongban=phongban
                               ) 
    else:
        id_yeucautuyendung = request.form.get("id")
        phongban = request.form.get("phongban")
        hoten = request.form.get("hoten")
        gioitinh = request.form.get("gioitinh")
        tuoi = request.form.get("tuoi")
        namkinhnghiem = request.form.get("namkinhnghiem")
        linkcv = request.files.get("linkcv")
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        save_path = os.path.join(FOLDER_CV,f"cv_{timestamp}.pdf")
        linkcv.save(save_path)
        if them_ungvientuyendung(id_yeucautuyendung,phongban,hoten,gioitinh,tuoi,namkinhnghiem,save_path):
            flash("Thêm ứng viên thành công")
        return redirect(f"muc2_2_1?id={id_yeucautuyendung}")
@app.route("/muc3_1", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def nhapthongtinlaodongmoi():
    
    if request.method == "GET":
        data= request.args.get("data")
        masothe = checkformatmst(int(laymasothemoi())+1)
        cacvitri= laycacvitri()
        cacto = laycacto()
        cacca = laycacca()
        data=request.args.get("scan-qrcode")
        datenow = datetime.now()
        macongty = current_user.macongty 
        return render_template("3_1.html", 
                                page="3.1 Nhập thông tin lao động mới",
                                qrcccd=data,
                                masothe=masothe,
                                ngaybatdau=datenow,
                                cacvitri=cacvitri,
                                cacto=cacto,
                                cacca=cacca,
                                macongty=macongty)
    elif request.method == "POST":
        try:
            anh = f"N'{request.form.get("anh")}'"
            masothe = f"'{request.form.get("masothe")}'"
            thechamcong = f"'{int(request.form.get("masothe"))}'"
            hoten = f"N'{request.form.get("hoten")}'"
            ngaysinh = f"'{request.form.get("ngaysinh")}'" if request.form.get("ngaysinh") else "NULL"
            gioitinh = f"N'{request.form.get("gioitinh")}'"
            cmt = f"'{request.form.get("cmt")}'"
            cccd = f"'{request.form.get("cccd")}'"
            ngaycapcccd = f"'{request.form.get("ngaycap")}'" if request.form.get("ngaycap") else "NULL"
            thuongtru = f"N'{request.form.get("thuongtru")}'" if request.form.get("thuongtru") else "NULL"
            noisinh = f"N'{request.form.get("noisinh")}'" if request.form.get("noisinh") else "NULL"
            tamtru = f"N'{request.form.get("tamtru")}'" if request.form.get("tamtru") else "NULL"
            quoctich = f"N'{request.form.get("quoctich")}'" if request.form.get("quoctich") else "NULL"
            dantoc = f"N'{request.form.get("dantoc")}'" if request.form.get("dantoc") else "NULL"
            tongiao = f"N'{request.form.get("tongiao")}'" if request.form.get("tongiao") else "NULL"
            hocvan = f"N'{request.form.get("hocvan")}'" if request.form.get("hocvan") else "NULL"
            thonxom = f"N'{request.form.get("thonxom")}'" if request.form.get("thonxom") else "NULL"
            phuongxa = f"N'{request.form.get("phuongxa")}'" if request.form.get("phuongxa") else "NULL"
            quanhuyen = f"N'{request.form.get("quanhuyen")}'" if request.form.get("quanhuyen") else "NULL"
            tinhthanhpho = f"N'{request.form.get("tinhthanhpho")}'" if request.form.get("tinhthanhpho") else "NULL"
            nganhang = f"N'{request.form.get("nganhang")}'" if request.form.get("nganhang") else "NULL"
            sotaikhoan = f"'{request.form.get("sotaikhoan")}'" if request.form.get("sotaikhoan") else "NULL"
            dienthoai = f"'{request.form.get("dienthoai")}'" if request.form.get("dienthoai") else "NULL"
            sobhxh = f"'{request.form.get("sobhxh")}'" if request.form.get("sobhxh") else 'NULL'
            masothue = f"'{request.form.get("masothue")}'" if request.form.get("masothue") else 'NULL'
            connho = f"N'{request.form.get("connho")}'" if request.form.get("connho") else 'NULL'
            tencon1 = f"N'{request.form.get("tenconnho1")}'" if request.form.get("tenconnho1") else 'NULL'
            ngaysinhcon1 = f"'{request.form.get("ngaysinhcon1")}'" if request.form.get("ngaysinhcon1") else 'NULL'
            tencon2 = f"N'{request.form.get("tenconnho2")}'" if request.form.get("tenconnho2") else 'NULL'
            ngaysinhcon2 = f"'{request.form.get("ngaysinhcon2")}'" if request.form.get("ngaysinhcon2") else 'NULL'
            tencon3 = f"N'{request.form.get("tenconnho3")}'" if request.form.get("tenconnho3") else 'NULL'
            ngaysinhcon3 = f"'{request.form.get("ngaysinhcon3")}'" if request.form.get("ngaysinhcon3") else 'NULL'
            tencon4 = f"N'{request.form.get("tenconnho4")}'" if request.form.get("tenconnho4") else 'NULL'
            ngaysinhcon4 = f"'{request.form.get("ngaysinhcon4")}'" if request.form.get("ngaysinhcon4") else 'NULL'
            tencon5 = f"N'{request.form.get("tenconnho5")}'" if request.form.get("tenconnho5") else 'NULL'
            ngaysinhcon5 = f"'{request.form.get("ngaysinhcon5")}'" if request.form.get("ngaysinhcon5") else 'NULL'
            jobdetailvn = f"N'{request.form.get("vitri")}'"
            line = f"'{request.form.get("line")}'"
            factory = f"'{current_user.macongty}'"
            hccategory = f"N'{request.form.get("hccategory")}'"
            gradecode = f"N'{request.form.get("gradecode")}'"
            department = f"N'{request.form.get("phongban")}'"
            chucvu = f"N'{request.form.get("chucvu")}'"
            employeetype = f"N'{request.form.get("loailaodong")}'"
            sectioncode = f"N'{request.form.get("mabophan")}'"
            sectiondescription = f"N'{request.form.get("bophan")}'"
            jobdetailen = f"N'{request.form.get("vitrien")}'"
            positioncode = f"N'{request.form.get("mavitri")}'"
            positioncodedescription = f"N'{request.form.get("tenvitri")}'"
            nguoithan = f"N'{request.form.get("nguoithan")}'" if request.form.get("nguoithan") else 'NULL'
            sdtnguoithan = f"N'{request.form.get("sdtnguoithan")}'" if request.form.get("sdtnguoithan") else 'NULL'
            luongcoban = 'NULL'
            tongphucap = 'NULL'
            kieuhopdong = 'NULL'

            ngaybatdauthuviec = 'NULL'
            ngayvao = 'GETDATE()'
            ngayketthuc = 'NULL'
            ngayketthucthuviec = 'NULL'
            ngaybatdauhdcthl1 = "NULL"
            ngayketthuchdcthl1 = "NULL"
            ngaybatdauhdcthl2 = "NULL"
            ngayketthuchdcthl2 = "NULL"
            ngaybatdauhdvth = "NULL"

            nhanvienmoi = f"({masothe},{thechamcong},{hoten},{dienthoai},{ngaysinh},{gioitinh},{cccd},{ngaycapcccd},N'Cục cảnh sát',{cmt},{thuongtru},{thonxom},{phuongxa},{quanhuyen},{tinhthanhpho},{dantoc},{quoctich},{tongiao},{hocvan},{noisinh},{tamtru},{sobhxh},{masothue},{nganhang},{sotaikhoan},{connho},{tencon1},{ngaysinhcon1},{tencon2},{ngaysinhcon2},{tencon3},{ngaysinhcon3},{tencon4},{ngaysinhcon4},{tencon5},{ngaysinhcon5},{anh},{nguoithan}, {sdtnguoithan},{kieuhopdong},{ngayvao},{ngayketthuc},{jobdetailvn},{hccategory},{gradecode},{factory},{department},{chucvu},{sectioncode},{sectiondescription},{line},{employeetype},{jobdetailen},{positioncode},{positioncodedescription},{luongcoban},N'Không',{tongphucap},{ngayvao},NULL,N'Đang làm việc',{ngayvao},'1',{ngaybatdauthuviec},{ngayketthucthuviec},{ngaybatdauhdcthl1},{ngayketthuchdcthl1},{ngaybatdauhdcthl2},{ngayketthuchdcthl2},{ngaybatdauhdvth},'N', '', GETDATE())"             
            ketqua = themnhanvienmoi(nhanvienmoi)
            if ketqua["ketqua"]:
                flash("Thêm lao động mới thành công !!!")
                ca = laycatheochuyen(request.form.get("line"))
                thangdangkycalamviec(
                    request.form.get("masothe"),
                    ca,
                    ca,
                    datetime.now().date().strftime("%Y-%m-%d"),  # This returns a datetime.date object
                    datetime(2054, 12, 31).date().strftime("%Y-%m-%d")  # Convert to datetime.date
                )
                flash(f"Tạo ca mặc định cho người mới là {ca} thành công !!!")  
                themtaikhoanmoi(
                    int(request.form.get("masothe")),
                    request.form.get("hoten"),
                    request.form.get("department"),
                    request.form.get("gradecode")
                )              
            else:
                flash("Thêm lao động mới thất bại !!!")
        except Exception as e:
            print(f"Them lao dong moi that bai: {e} !!!")
        finally:
            return redirect("/muc3_1")
        
@app.route("/muc3_2", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def thaydoithongtinlaodong():
    
    if request.method == "GET":
        return render_template("3_2.html", page="3.2 Thay đổi thông tin người lao động")
    else:
        try:
            trangthailamviec = request.form.get("trangthai")
            thechamcong = request.form.get("thechamcong")
            cccd = request.form.get("cccd")
            ngaycapcccd = request.form.get("ngaycapcccd")
            hoten = request.form.get("hoten")
            ngaysinh = request.form.get("ngaysinh")
            gioitinh = request.form.get("gioitinh")
            cmt = request.form.get("cmt")
            quoctich = request.form.get("quoctich")
            dienthoai = request.form.get("dienthoai")
            thonxom = request.form.get("thonxom")
            phuongxa = request.form.get("phuongxa")
            quanhuyen = request.form.get("quanhuyen")
            tinhthanhpho = request.form.get("tinhthanhpho")
            thuongtru = request.form.get("thuongtru")
            tamtru = request.form.get("tamtru")
            noisinh = request.form.get("noisinh")
            dantoc = request.form.get("dantoc")
            tongiao = request.form.get("tongiao")
            hocvan = request.form.get("hocvan")
            masothue = request.form.get("masothue")
            nganhang = request.form.get("nganhang")
            sotaikhoan = request.form.get("sotaikhoan")
            connho = request.form.get("connho")
            mst = request.form.get("mst")
            tenconnho1 = request.form.get("tenconnho1")
            tenconnho2 = request.form.get("tenconnho2")
            tenconnho3 = request.form.get("tenconnho3")
            tenconnho4 = request.form.get("tenconnho4")
            tenconnho5 = request.form.get("tenconnho5")
            ngaysinhcon1 = request.form.get("ngaysinhcon1")
            ngaysinhcon2 = request.form.get("ngaysinhcon2")
            ngaysinhcon3 = request.form.get("ngaysinhcon3")
            ngaysinhcon4 = request.form.get("ngaysinhcon4")
            ngaysinhcon5 = request.form.get("ngaysinhcon5")
            jobtitlevn = request.form.get("jobtitlevn") 
            jobtitleen = request.form.get("jobtitleen") 
            positioncode = request.form.get("positioncode")
            positioncodedescription = request.form.get("positioncodedescription")
            chucvu =  request.form.get("chucvu")
            line = request.form.get("line")
            department = request.form.get("department")
            sectioncode = request.form.get("sectioncode")
            sectiondescription = request.form.get("sectiondescription")
            hccategory = request.form.get("hccategory")
            employeetype = request.form.get("employeetype")
            gradecode = request.form.get("gradecode")
            factory = request.form.get("factory")
            kieuhopdong = request.form.get("kieuhopdong")
            ngaybatdau = request.form.get("ngaybatdau")
            ngayketthuc = request.form.get("ngayketthuc")
            mucluong = request.form.get("mucluong").replace(',','')
            phucap = request.form.get("phucap").replace(',','')
            tienphucap = request.form.get("tienphucap")
            nguoithan = request.form.get("nguoithan")
            sdtnguoithan = request.form.get("sdtnguoithan")
            ngayvao = request.form.get("ngayvao")
            ngaynghi = request.form.get("ngaynghi")
            ngaykyhdtv = request.form.get("ngaykyhdtv")
            ngayhethanhdtv = request.form.get("ngayhethanhdtv")
            
            query = f"UPDATE HR.dbo.Danh_sach_CBCNV SET "
            if thechamcong:
                query += f"The_cham_cong = '{thechamcong}',"
            else:
                query += f"The_cham_cong = NULL,"
                
            if cccd: 
                query += f"CCCD = '{cccd}',"
            else:
                query += f"CCCD = NULL,"
                
            if ngaycapcccd: 
                query += f"Ngay_cap = '{ngaycapcccd}',"
            else:
                query += f"Ngay_cap = NULL,"
                
            if hoten: 
                query += f"Ho_ten = N'{hoten}',"
            else:
                query += f"Ho_ten = NULL,"
                
            if ngaysinh: 
                query += f"Ngay_sinh = '{ngaysinh}',"
            else:
                query += f"Ngay_sinh = NULL,"
                
            if gioitinh: 
                query += f"Gioi_tinh = N'{gioitinh}',"
            else:
                query += f"Gioi_tinh = NULL,"
                    
            if cmt: 
                query += f"CMT = N'{cmt}',"
            else:
                query += f"CMT = NULL,"   
                
            if quoctich: 
                query += f"Quoc_tich = N'{quoctich}',"
            else:
                query += f"Quoc_tich = NULL," 
                
            if dienthoai: 
                query += f"Sdt = N'{dienthoai}',"
            else:
                query += f"Sdt = NULL,"
                
            if thonxom: 
                query += f"Thon_xom = N'{thonxom}',"
            else:
                query += f"Thon_xom = NULL,"
                
            if phuongxa: 
                query += f"Phuong_xa = N'{phuongxa}',"
                
            if quanhuyen: 
                query += f"Quan_huyen = N'{quanhuyen}',"
            else:
                query += f"Quan_huyen = NULL,"
                
            if tinhthanhpho: 
                query += f"Tinh_TP = N'{tinhthanhpho}',"
            else:
                query += f"Tinh_TP = NULL,"
            
            if thuongtru:
                query += f"Dia_chi_thuong_tru = N'{thuongtru}',"
            else:
                query += f"Dia_chi_thuong_tru = NULL,"
                
            if tamtru:
                query += f"Dia_chi_tam_tru = N'{tamtru}',"
            else:
                query += f"Dia_chi_tam_tru = NULL,"
                
            if noisinh:
                query += f"Noi_sinh = N'{noisinh}',"
            else:
                query += f"Noi_sinh = NULL,"
            
            if dantoc: 
                query += f"Dan_toc = N'{dantoc}',"
            else:
                query += f"Dan_toc = NULL,"
                    
            if tongiao: 
                query += f"Ton_giao = N'{tongiao}',"
            else:
                query += f"Ton_giao = NULL,"
                    
            if hocvan: 
                query += f"Trinh_do = N'{hocvan}',"
            else:
                query += f"Trinh_do = NULL,"
                
            if masothue: 
                query += f"Ma_so_thue = N'{masothue}',"
            else:
                query += f"Ma_so_thue = NULL,"   
                
            if nganhang: 
                query += f"Ngan_hang = N'{nganhang}',"
            else:
                query += f"Ngan_hang = NULL,"   
                
            if sotaikhoan: 
                query += f"So_tai_khoan = N'{sotaikhoan}',"
            else:
                query += f"So_tai_khoan = NULL,"
                
            if connho: 
                query += f"Con_nho = N'{connho}',"
            else:
                query += f"Con_nho = NULL,"

            if tenconnho1: 
                query += f"Ten_con_nho_1 = N'{tenconnho1}',"
            else:
                query += f"Ten_con_nho_1 = NULL,"
            
            if tenconnho2: 
                query += f"Ten_con_nho_2 = N'{tenconnho2}',"
            else:
                query += f"Ten_con_nho_2 = NULL,"
                
            if tenconnho3: 
                query += f"Ten_con_nho_3 = N'{tenconnho3}',"
            else:
                query += f"Ten_con_nho_3 = NULL,"
                
            if tenconnho4: 
                query += f"Ten_con_nho_4 = N'{tenconnho4}',"
            else:
                query += f"Ten_con_nho_4 = NULL,"
            
            if tenconnho5: 
                query += f"Ten_con_nho_5 = N'{tenconnho5}',"
            else:
                query += f"Ten_con_nho_5 = NULL,"
            
            if ngaysinhcon1: 
                query += f"Ngay_sinh_con_nho_1 = '{ngaysinhcon1}',"
            else:
                query += f"Ngay_sinh_con_nho_1 = NULL,"
            
            if ngaysinhcon2: 
                query += f"Ngay_sinh_con_nho_2 = '{ngaysinhcon2}',"
            else:
                query += f"Ngay_sinh_con_nho_2 = NULL,"
                
            if ngaysinhcon3: 
                query += f"Ngay_sinh_con_nho_3 = '{ngaysinhcon3}',"
            else:
                query += f"Ngay_sinh_con_nho_3 = NULL,"
                
            if ngaysinhcon4: 
                query += f"Ngay_sinh_con_nho_4 = '{ngaysinhcon4}',"
            else:
                query += f"Ngay_sinh_con_nho_4 = NULL,"
                
            if ngaysinhcon5: 
                query += f"Ngay_sinh_con_nho_5 = '{ngaysinhcon5}',"
            else:
                query += f"Ngay_sinh_con_nho_5 = NULL,"
            
            if nguoithan:
                query += f"Nguoi_than = N'{nguoithan}',"
            else:
                query += f"Nguoi_than = NULL," 
                
            if sdtnguoithan:
                query += f"Sdt_Nguoithan = '{sdtnguoithan}',"
            else:
                query += f"Sdt_Nguoithan = NULL," 
            
            if jobtitlevn: 
                query += f"Job_title_VN = N'{jobtitlevn}',"
            else:
                query += f"Job_title_VN = NULL,"
                
            if jobtitleen: 
                query += f"Job_title_EN = '{jobtitleen}',"
            else:
                query += f"Job_title_EN = NULL,"
                
            if positioncode: 
                query += f"Position_code = '{positioncode}',"
            else:
                query += f"Position_code = NULL,"
            
            if positioncodedescription: 
                query += f"Position_code_description = '{positioncodedescription}',"
            else:
                query += f"Position_code_description = NULL,"  
            
            if chucvu: 
                query += f"Chuc_vu = N'{chucvu}',"
            else:
                query += f"Chuc_vu = NULL," 
                
            if line: 
                query += f"Line = '{line}',"
            else:
                query += f"Line = NULL,"     
            
            if department: 
                query += f"Department = '{department}',"
            else:
                query += f"Department = NULL,"  
                
            if sectioncode: 
                query += f"Section_code = '{sectioncode}',"
            else:
                query += f"Section_code = NULL,"     
            
            if sectiondescription: 
                query += f"Section_description = '{sectiondescription}',"
            else:
                query += f"Section_description = NULL," 
                
            if hccategory: 
                query += f"Headcount_category = '{hccategory}',"
            else:
                query += f"Headcount_category = NULL," 
                
            if employeetype: 
                query += f"Emp_type = '{employeetype}',"
            else:
                query += f"Emp_type = NULL," 
                
            if gradecode: 
                query += f"Grade_code = '{gradecode}',"
            else:
                query += f"Grade_code = NULL," 
                
            if factory: 
                query += f"Factory = '{factory}',"
            else:
                query += f"Factory = NULL,"
                
            if kieuhopdong:
                query += f"Loai_hop_dong = N'{kieuhopdong}',"
            else:
                query += f"Loai_hop_dong = NULL,"
                
            if ngaybatdau:
                query += f"Ngay_ky_HD = '{ngaybatdau}',"
            else:
                query += f"Ngay_ky_HD = NULL,"
            
            if ngayketthuc:
                query += f"Ngay_het_han_HD = '{ngayketthuc}',"
            else:
                query += f"Ngay_het_han_HD = NULL,"
                
            if mucluong:
                query += f"Luong_co_ban = '{mucluong}',"
            else:
                query += f"Luong_co_ban = NULL,"
                
            if phucap:
                query += f"Phu_cap = N'{phucap}',"
            else:
                query += f"Phu_cap = NULL,"
                
            if tienphucap:
                query += f"Tong_phu_cap = '{tienphucap}',"
            else:
                query += f"Tong_phu_cap = NULL,"
            if trangthailamviec:
                query += f"Trang_thai_lam_viec = N'{trangthailamviec}',"
            else:
                query += f"Trang_thai_lam_viec = NULL,"
            if ngayvao:
                query += f"Ngay_vao = '{ngayvao}',"
            else:
                query += f"Ngay_vao = NULL,"
            if ngaynghi:
                query += f"Ngay_nghi = '{ngaynghi}',"
            else:
                query += f"Ngay_nghi = NULL,"
            if ngaykyhdtv:
                query += f"Ngay_ky_HDTV = '{ngaykyhdtv}',"
            else:
                query += f"Ngay_ky_HDTV = NULL,"
            if ngayhethanhdtv:
                query += f"Ngay_het_han_HDTV = '{ngayhethanhdtv}',"
            else:
                query += f"Ngay_het_han_HDTV = NULL,"
            query = query[:-1] + f" WHERE MST = '{mst}' AND Factory='{current_user.macongty}'"
            conn = pyodbc.connect(url_database_pyodbc)
            cursor = conn.cursor()
            cursor.execute(query)
            conn.commit()
            conn.close()
            flash("Cập nhật thông tin người lao động thành công !!!")
        except Exception as e:
            flash(f"Cập nhật thông tin người lao động thất bại: {e} !!!")
        return redirect("/muc3_2")
    
@app.route("/muc3_3", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def inhopdonglaodong():
    try:
        if request.method == "GET":
            mst = request.args.get("mst")
            if mst:
                danhsach = laydanhsach_hopdong_theomst(mst)
            else:
                danhsach = []
            return render_template("3_3.html", page="3.3 Quản lý hợp đồng lao động",danhsach=danhsach)
        elif request.method == "POST":
            nhamay = current_user.macongty
            mst = request.form.get("form_manhanvien")
            hoten = request.form.get("form_hovaten")
            gioitinh = request.form.get("form_gioitinh")
            ngaysinh =  request.form.get("form_ngaysinh")
            thuongtru = request.form.get("form_thuongtru")
            tamtru = request.form.get("form_tamtru")
            cccd = request.form.get("form_cccd")
            ngaycapcccd = request.form.get("form_ngaycapcccd")
            capbac =  request.form.get("gradecode")
            loaihopdong = request.form.get("form_loaihopdong")
            chucdanh = request.form.get("chucdanh")
            phongban = request.form.get("department")
            chuyen = request.form.get("chuyen")
            luongcoban = request.form.get("luongcoban")
            phucap = request.form.get("phucap")
            ngaybatdau = request.form.get("form_ngaykyhopdong")
            ngayketthuc = request.form.get("form_ngayhethanhopdong")
            vitrien = request.form.get("vitrien")
            employeetype = request.form.get("employeetype")
            positioncode = request.form.get("positioncode")
            postitioncodedescription = request.form.get("postitioncodedescription")
            hccategory = request.form.get("hccategory")
            sectioncode = request.form.get("sectioncode")
            sectiondescription = request.form.get("sectiondescription")
            if themhopdongmoi(nhamay,mst,hoten,gioitinh,ngaysinh,thuongtru,tamtru,cccd,ngaycapcccd,capbac,loaihopdong,chucdanh,phongban,chuyen,luongcoban,phucap,ngaybatdau,ngayketthuc):
                print("Thêm hợp đồng thành công !!!")
                # capnhatthongtinhopdong(nhamay,mst,loaihopdong,chucdanh,chuyen,luongcoban,phucap,ngaybatdau,ngayketthuc,vitrien,employeetype,positioncode,postitioncodedescription,hccategory,sectioncode,sectiondescription)
            else:
                print("Thêm hợp đồng thất bại")
            return redirect("/muc3_3")
    except:
        return redirect("/muc3_3")
    
@app.route("/muc3_4", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def danhsachsaphethanhopdong():
    if request.method == "GET":
        danhsach = laydanhsachsaphethanhopdong()
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("/3_4.html",
                               danhsach=paginated_rows,
                               pagination=pagination,
                               total = total)
    if request.method == "POST":
        danhsach = laydanhsachsaphethanhopdong()
        result = []
        for user in danhsach:
            result.append({
                "MST": user[0],
            "Thẻ chấm công": user[1],
            "Họ tên": user[2],
            "Số điện thoại": user[3],
            "Ngày sinh": datetime.strptime(user[4], '%Y-%m-%d').strftime("%d/%m/%Y") if user[4] else None,
            "Giới tính": user[5],
            "CCCD": user[6],
            "Ngày cấp CCCD": datetime.strptime(user[7], '%Y-%m-%d').strftime("%d/%m/%Y") if user[7] else None ,
            "Nơi cấp": user[8],
            "CMT": user[9],
            "Thường trú": user[10],
            "Thôn xóm": user[11],
            "Phường xã": user[12],
            "Quận huyện": user[13],
            "Tỉnh thành phố": user[14],
            "Dân tộc": user[15],
            "Quốc tịch": user[16],
            "Tôn giáo": user[17],
            "Học vấn": user[18],
            "Nơi sinh": user[19],
            "Tạm trú": user[20],
            "Số BHXH": user[21],
            "Mã số thuế": user[22],
            "Ngân hàng": user[23],
            "Số tài khoản": user[24],
            "Con nhỏ": user[25],
            "Tên con 1": user[26],
            "Ngày sinh con 1": user[27],
            "Tên con 2": user[28],
            "Ngày sinh con 2": user[29],
            "Tên con 3": user[30],
            "Ngày sinh con 3": user[31],
            "Tên con 4": user[32],
            "Ngày sinh con 4": user[33],
            "Tên con 5": user[34],
            "Ngày sinh con 5": user[35],
            "Ảnh chân dung": user[36],
            "Người thân": user[37],
            "SĐT liên hệ": user[38],
            "Loại hợp đồng": user[39],
            "Ngày ký HĐ": datetime.strptime(user[40], '%Y-%m-%d').strftime("%d/%m/%Y") if user[40] else None,
            "Ngày hết hạn": datetime.strptime(user[41], '%Y-%m-%d').strftime("%d/%m/%Y") if user[41] else None,
            "Job title VN": user[42],
            "HC category": user[43],
            "Gradecode": user[44],
            "Factory": user[45],
            "Department": user[46],
            "Chức vụ": user[47],
            "Section code": user[48],
            "Section description": user[49],
            "Line": user[50],
            "Employee type": user[51],
            "Job title EN": user[52],
            "Position code": user[53],
            "Position description": user[54],
            "Lương cơ bản": user[55],
            "Phụ cấp": user[56],
            "Tiền phụ cấp": user[57],
            "Ngày vào": datetime.strptime(user[58], '%Y-%m-%d').strftime("%d/%m/%Y"),
            "Ngày nghỉ": datetime.strptime(user[59], '%Y-%m-%d').strftime("%d/%m/%Y") if user[59] else None,
            "Trạng thái": user[60],
            "Ngày vào nối thâm niên": datetime.strptime(user[61], '%Y-%m-%d').strftime("%d/%m/%Y") if user[61] else None,
            "Mật khẩu": user[62],
            "Ngày kí HĐ Thử việc": datetime.strptime(user[63], '%Y-%m-%d').strftime("%d/%m/%Y") if user[63] else None,
            "Ngày hết hạn HĐ Thử việc": datetime.strptime(user[64], '%Y-%m-%d').strftime("%d/%m/%Y") if user[64] else None,
            "Ngày kí HĐ xác định thời hạn lần 1": datetime.strptime(user[65], '%Y-%m-%d').strftime("%d/%m/%Y") if user[65] else None,
            "Ngày hết hạn HĐ xác định thời hạn lần 1": datetime.strptime(user[66], '%Y-%m-%d').strftime("%d/%m/%Y") if user[66] else None,
            "Ngày kí HĐ HĐ xác định thời hạn lần 2": datetime.strptime(user[67], '%Y-%m-%d').strftime("%d/%m/%Y") if user[67] else None,
            "Ngày hết hạn HĐ xác định thời hạn lần 2": datetime.strptime(user[68], '%Y-%m-%d').strftime("%d/%m/%Y") if user[68] else None,
            "Ngày kí HĐ không thời hạn": datetime.strptime(user[69], '%Y-%m-%d').strftime("%d/%m/%Y") if user[69] else None,
            "Ghi chú": user[71] if user[71] else None
            })
        df = pd.DataFrame(result)
        thoigian = datetime.now().strftime("%d%m%Y%H%M%S")
        df.to_excel(os.path.join(FOLDER_XUAT, f"saphethan_{thoigian}.xlsx"), index=False)
        print("Tải file thành công !!!")
        return send_file(os.path.join(FOLDER_XUAT, f"saphethan_{thoigian}.xlsx"), as_attachment=True)

@app.route("/muc5_1_1", methods=["GET","POST"])
@login_required
@roles_required('sa','tbp','gd')
def nhapkpi():
    if request.method == "GET":
        danhsach = laydanhsachkpichuaduyet(current_user.masothe,current_user.macongty)
        return render_template("5_1_1.html",page="Upload KPI",danhsach=danhsach)
    if request.method == "POST":
        try:
            file = request.files['file']
            if file:
                ngaylam = datetime.now().strftime("%d%m%Y%H%M%S")
                filepath = os.path.join(FOLDER_NHAP, f"kpi_{current_user.masothe}_{ngaylam}.xlsx")
                file.save(filepath)
                data = pd.read_excel(filepath).to_dict(orient="records")
                delete_kpidata(current_user.masothe,current_user.macongty)
                for row in data[1:]:
                    values=[]
                    for x in row.items():
                        values.append(str(x[1]).replace("'","") if x[1] else "")
                    if not insert_kpidata(current_user.masothe,current_user.macongty,values):
                        print("Upload new KPI failed: Cannot insert data !!!")
                        return redirect("/muc5_1_1")
                guimailthongbaodaguikpi(current_user.macongty,current_user.masothe,current_user.hoten)
                print("Upload new KPI successfully !!!")
            else:
                print("Upload new KPI failed: Cannot found data !!!")
        except Exception as e:
            print(f"Upload new KPI failed {e} !!!")
            print("Upload new KPI failed !!!")
        return redirect("/muc5_1_1")

@app.route("/muc5_1_2", methods=["GET","POST"])
@login_required
@roles_required('sa','gd')
def duyetkpi():
    if request.method == "GET":
        congty = request.args.get("company")
        pic = request.args.get("pic")
        mst = pic.split("_")[0] if pic else None
        danhsachquanly = laydanhsachquanly(congty)
        danhsach = laydanhsachkpichuaduyet(mst,congty)
        return render_template("5_1_2.html",page="Approve KPI",danhsach=danhsach,danhsachquanly=danhsachquanly)
    if request.method == "POST":
        congty = request.form.get("company")
        pic = request.args.get("pic")
        mst = pic.split("_")[0] if pic else None
        hoten = pic.split("_")[1] if pic else None
        email = layemailquanly(congty,mst)
        pheduyet = request.form.get("pheduyet")
        if pheduyet == "co":
            pheduyetkpi(mst,congty)
            guimailthongbaodapheduyetkpi(congty,mst,hoten,email)
        else:
            tuchoikpi(mst,congty)
            guimailthongbaodatuchoikpi(congty,mst,hoten,email)
        return redirect(f"/muc5_1_2?company={congty}&mst={mst}")

@app.route("/muc5_1_3_1", methods=["GET","POST"])
@login_required
@roles_required('sa','gd','tbp')
def baocaocanam():
    if request.method == "GET":
        congty = request.args.get("company")
        mst = request.args.get("mst")
        danhsachquanly = laydanhsachquanly(congty)
        danhsach = laydanhsachkpidaduyet(mst,congty)
        return render_template("5_1_3_1.html",page="Performance Report All Year",danhsach=danhsach,danhsachquanly=danhsachquanly)

@app.route("/muc5_1_3_2", methods=["GET","POST"])
@login_required
@roles_required('sa','gd','tbp')
def baocaoytd():
    congty = request.args.get("company")
    mst = request.args.get("mst")
    danhsachquanly = laydanhsachquanly(congty)
    danhsach = laydanhsachkpidaduyet(mst,congty)
    return render_template("5_1_3_2.html",page="Performance Report Year to date",danhsach=danhsach,danhsachquanly=danhsachquanly)


    
@app.route("/muc6_1", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def dieuchuyen():
    try:
        if request.method == "POST":
            mst = request.form["mst"]
            loaidieuchuyen = request.form["loaidieuchuyen"]
            ngaydieuchuyen = request.form.get("ngaydieuchuyen")
            ghichu = request.form.get("ghichu")
            
            vitricu = request.form.get("vitricu")
            vitrimoi = request.form.get("vitrimoi")
            
            vitriencu = request.form.get("vitriencu")
            vitrienmoi = request.form.get("vitrienmoi")
            
            chuyencu = request.form.get("chuyencu")
            chuyenmoi = request.form.get("chuyenmoi")
            
            gradecodecu = request.form.get("gradecodecu")
            gradecodemoi = request.form.get("gradecodemoi")
            
            sectioncodecu = request.form.get("sectioncodecu")
            sectioncodemoi = request.form.get("sectioncodemoi")
            
            hccategorycu = request.form.get("hccategorycu")
            hccategorymoi = request.form.get("hccategorymoi")
            
            departmentcu = request.form.get("departmentcu")
            departmentmoi = request.form.get("departmentmoi")
            
            sectiondescriptioncu = request.form.get("sectiondescriptioncu")
            sectiondescriptionmoi = request.form.get("sectiondescriptionmoi")
            
            employeetypecu = request.form.get("employeetypecu") 
            employeetypemoi = request.form.get("employeetypemoi")
            
            positioncodecu = request.form.get("positioncodecu") 
            positioncodemoi = request.form.get("positioncodemoi") 
            
            positioncodedescriptioncu = request.form.get("positioncodedescriptioncu") 
            positioncodedescriptionmoi = request.form.get("positioncodedescriptionmoi") 
            
            khongdoica = request.form.get("khongdoica") 
            
            if loaidieuchuyen == "Chuyển vị trí":
                try:
                    ketqua = dieuchuyennhansu(mst,
                                    loaidieuchuyen,
                                    vitricu,
                                    vitrimoi,
                                    chuyencu,
                                    chuyenmoi,
                                    gradecodecu,
                                    gradecodemoi,
                                    sectioncodecu,
                                    sectioncodemoi,
                                    hccategorycu,
                                    hccategorymoi,
                                    departmentcu,
                                    departmentmoi,
                                    sectiondescriptioncu,
                                    sectiondescriptionmoi,
                                    employeetypecu,
                                    employeetypemoi,
                                    positioncodedescriptioncu,
                                    positioncodedescriptionmoi,
                                    positioncodecu,
                                    positioncodemoi,
                                    vitriencu,
                                    vitrienmoi,
                                    ngaydieuchuyen,
                                    ghichu,
                                    khongdoica
                                    )
                    if ketqua["ketqua"]:
                        flash("Điều chuyển thành công !!!")
                    else:
                        flash(f"Điều chuyển thất bại, lí do: {ketqua["lido"]}, query: {ketqua["query"]} !!!")
                except Exception as e:
                    flash(f"Điều chuyển thất bại, lí do: {e}")
                return redirect(f"/muc6_1")
                
            elif loaidieuchuyen == "Nghỉ việc":
                try:
                    ketqua = dichuyennghiviec(mst,
                        vitricu,
                        chuyencu,
                        gradecodecu,
                        hccategorycu,
                        ngaydieuchuyen,
                        ghichu)
                    if ketqua["ketqua"]:
                        flash("Điều chuyển thành công !!!")
                    else:
                        flash(f"Điều chuyển thất bại, lí do: {ketqua["lido"]}, query: {ketqua["query"]} !!!")
                except Exception as e:
                    flash(f"Điều chuyển thất bại, lí do: {e}")
                return redirect(f"/muc6_1")
            elif loaidieuchuyen=="Nghỉ thai sản":
                try:
                    ketqua = dichuyennghithaisan(mst,
                                vitricu,
                                chuyencu,
                                gradecodecu,
                                hccategorycu,
                                ngaydieuchuyen
                                )
                    if ketqua["ketqua"]:
                        flash("Điều chuyển thành công !!!")
                    else:
                        flash(f"Điều chuyển thất bại, lí do: {ketqua["lido"]}, query: {ketqua["query"]} !!!")
                except Exception as e:
                    flash(f"Điều chuyển thất bại, lí do: {e}")
                return redirect(f"/muc6_1")
            elif loaidieuchuyen=="Thai sản đi làm lại":
                try:
                    ketqua = dichuyenthaisandilamlai(mst,
                                    vitricu,
                                    vitrimoi,
                                    chuyencu,
                                    chuyenmoi,
                                    gradecodecu,
                                    gradecodemoi,
                                    sectioncodecu,
                                    sectioncodemoi,
                                    hccategorycu,
                                    hccategorymoi,
                                    departmentcu,
                                    departmentmoi,
                                    sectiondescriptioncu,
                                    sectiondescriptionmoi,
                                    employeetypecu,
                                    employeetypemoi,
                                    positioncodedescriptioncu,
                                    positioncodedescriptionmoi,
                                    positioncodecu,
                                    positioncodemoi,
                                    vitriencu,
                                    vitrienmoi,
                                    ngaydieuchuyen
                            )
                    if ketqua["ketqua"]:
                        flash("Điều chuyển thành công !!!")
                    else:
                        flash(f"Điều chuyển thất bại, lí do: {ketqua["lido"]}, query: {ketqua["query"]} !!!")
                except Exception as e:
                    flash(f"Điều chuyển thất bại, lí do: {e}")
                return redirect(f"/muc6_1")
            return redirect(f"/muc6_1")
        elif request.method == "GET":
            cacvitri= laycacvitri()
            return render_template("6_1.html",
                            cacvitri=cacvitri,
                            page="6.1 Điều chuyển chức vụ, bộ phận")
    except Exception as e:
        flash(e)
        cacvitri= laycacvitri()
        return render_template("6_1.html",
                            cacvitri=cacvitri,
                            page="6.1 Điều chuyển chức vụ, bộ phận")
    
@app.route("/muc6_2", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def lichsudieuchuyen():
    
    if request.method == "GET":
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        ngay = request.args.get("ngay")
        kieudieuchuyen = request.args.get("kieudieuchuyen")
        rows = laylichsucongtac(mst,hoten,ngay,kieudieuchuyen)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("6_2.html", page="6.2 Lịch sử điều chuyển",
                               danhsach=paginated_rows, 
                               pagination=pagination,
                               mst=mst, 
                               count=count)
    if request.method == "POST":
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        ngay = request.args.get("ngay")
        kieudieuchuyen = request.args.get("kieudieuchuyen")
        data = laylichsucongtac(mst,hoten,ngay,kieudieuchuyen)
        df = DataFrame(data)
        df["Ngày thực hiện"] = to_datetime(df['Ngày thực hiện'])
        df["Ngày chính thức"] = to_datetime(df['Ngày chính thức'])
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook.active
        # Create a date format for short date
        date_format = NamedStyle(name="short_date", number_format="DD/MM/YYYY")
        if "short_date" not in workbook.named_styles:
            workbook.add_named_style(date_format)
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                    if cell.column_letter == 'C' and cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY'
                    if cell.column_letter == 'K' and cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY'
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=dieuchuyen_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

@app.route("/muc6_3", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def lichsucongviec():
    if request.method == "GET":
        mst = request.args.get("mst")
        chuyen = request.args.get("chuyen")
        bophan = request.args.get("bophan")
        rows = laylichsucongviec(mst,chuyen,bophan)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("6_3.html", page="6.3 Lịch sử công việc",
                               danhsach=paginated_rows, 
                               pagination=pagination,
                               count=count)
        
    elif request.method == "POST":
        mst = request.form.get("mst")
        chuyen = request.form.get("chuyen")
        bophan = request.form.get("bophan")
        rows = laylichsucongviec(mst,chuyen,bophan)
        data = [{
            "Mã công ty": row[0],
            "Mã số thẻ": row[1],
            "Họ tên": row[2],
            "Chuyền": row[3],
            "Bộ phận": row[4],
            "Chức danh": row[5],
            "Cấp bậc": row[6],
            "HC category": row[11],
            "Trạng thái": row[7],
            "Ngày bắt đầu": row[8],
            "Ngày kết thúc": row[9]
        } for row in rows]
        df = DataFrame(data)
        df["Ngày bắt đầu"] = to_datetime(df['Ngày bắt đầu'])
        df["Ngày kết thúc"] = to_datetime(df['Ngày kết thúc'])
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook.active
        # Create a date format for short date
        date_format = NamedStyle(name="short_date", number_format="DD/MM/YYYY")
        if "short_date" not in workbook.named_styles:
            workbook.add_named_style(date_format)
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    # Apply the date format to column L (assuming 'Ngày thực hiện' is in column 'L')
                    if cell.column_letter == 'J' and cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY'
                    if cell.column_letter == 'K' and cell.value is not None:
                        cell.number_format = 'DD/MM/YYYY'
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=lichsu_congviec_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
    
@app.route("/muc7_1_1", methods=["GET","POST"]) # Đổi ca làm việc
@login_required
@roles_required('hr','sa','gd')
def khaibaochamcong():
    if request.method == "GET":
        try:
            mst = request.args.get("mst")
            chuyen = request.args.get("chuyen") 
            phongban = request.args.get("phongban") 
            rows = laydanhsachcahientai(mst,chuyen,phongban)
            count = len(rows)
            current_page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 15
            total = len(rows)
            start = (current_page - 1) * per_page
            end = start + per_page
            paginated_rows = rows[start:end]
            pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
            cacca = laycacca()
            return render_template("7_1_1.html",
                                    page="7.1.1 Đổi ca làm việc",
                                    danhsach=paginated_rows,
                                    pagination=pagination,
                                    count=count,
                                    cacca=cacca)
        except:
            return render_template("7_1_1.html",
                                    page="7.1.1 Đổi ca làm việc",
                                    danhsach=[])
    elif request.method == "POST":
        mst = request.form.get("mst")
        chuyen = request.form.get("chuyen") 
        phongban = request.form.get("phongban") 
        rows = laydanhsachcahientai(mst,chuyen,phongban)
        data =[]
        for row in rows:
            data.append({
                "Nhà máy": row[0],
                "Mã số thẻ": row[1],
                "Họ tên": row[2],
                "Chuyền tổ": row[3], 
                "Phòng ban": row[4],
                "Ca": row[5],
                "Đổi từ ngày": row[6],
                "Đổi đến ngày": row[7]
            })
        df = pd.DataFrame(data)
        thoigian = datetime.now().strftime("%d%m%Y%H%M%S")
        df.to_excel(os.path.join(FOLDER_XUAT, f"doica_{thoigian}.xlsx"), index=False)
        print("Tải file thành công !!!")
        return send_file(os.path.join(FOLDER_XUAT, f"doica_{thoigian}.xlsx"), as_attachment=True)
            
@app.route("/muc7_1_2", methods=["GET","POST"]) # Danh sách lỗi chấm công
@login_required
def loichamcong():
    mstthuky = request.args.get("mstthuky")
    mst = request.args.get("mst")
    chuyen = request.args.get("chuyen")
    bophan = request.args.get("bophan")
    ngay = request.args.get("ngay")
    danhsach = laydanhsachloithe(mst,chuyen,bophan,ngay,mstthuky)
    count = len(danhsach)
    current_page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10
    total = len(danhsach)
    start = (current_page - 1) * per_page
    end = start + per_page
    paginated_rows = danhsach[start:end]
    pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
    return render_template("7_1_2.html",
                            page="Lỗi chấm công",
                            danhsach=paginated_rows, 
                            pagination=pagination,
                            count=count)


@app.route("/muc7_1_3", methods=["GET"]) # Danh sách điểm danh bù
@login_required
def diemdanhbu():
    mstthuky = request.args.get("mstthuky")
    mstquanly = request.args.get("mstquanly")
    mst = request.args.get("mst")
    hoten = request.args.get("hoten")
    chucvu = request.args.get("chucvu")
    chuyen = request.args.get("chuyen")
    bophan = request.args.get("bophan")
    loaidiemdanh = request.args.get("loaidiemdanh")
    ngay = request.args.get("ngay")
    lido = request.args.get("lido")
    trangthai = request.args.get("trangthai")
    danh_sach_chuyen = laydanhsachchuyen()
    danh_sach_bophan = laydanhsachbophan()
    danhsach = laydanhsachdiemdanhbu(mst,hoten,chucvu,chuyen,bophan,loaidiemdanh,ngay,lido,trangthai,mstquanly,mstthuky)
    count = len(danhsach)
    current_page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10
    total = len(danhsach)
    start = (current_page - 1) * per_page
    end = start + per_page
    paginated_rows = danhsach[start:end]
    pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
    return render_template("7_1_3.html",
                        page="Lỗi chấm công",
                        danhsach=paginated_rows, 
                        pagination=pagination,
                        count=count,
                        danh_sach_chuyen=danh_sach_chuyen,
                        danh_sach_bophan=danh_sach_bophan)
 
@app.route("/muc7_1_4", methods=["GET"]) # Danh sách xin nghỉ phép 
@login_required
def xinnghiphep():
    mstthuky = request.args.get("mstthuky")
    mstquanly = request.args.get("mstquanly")
    mst = request.args.get("mst")
    hoten = request.args.get("hoten")
    chucvu = request.args.get("chucvu")
    chuyen = request.args.get("chuyen")
    bophan = request.args.get("bophan")
    ngay = request.args.get("ngaynghi")
    lydo = request.args.get("lydo")
    trangthai = request.args.get("trangthai")
    danhsach = laydanhsachxinnghiphep(mst,hoten,chucvu,chuyen,bophan,ngay,lydo,trangthai,mstquanly,mstthuky)
    count = len(danhsach)
    current_page = request.args.get(get_page_parameter(), type=int, default=1)
    per_page = 10
    total = len(danhsach)
    start = (current_page - 1) * per_page
    end = start + per_page
    paginated_rows = danhsach[start:end]
    pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
    return render_template("7_1_4.html",
                        page="Lỗi chấm công",
                        danhsach=paginated_rows, 
                        pagination=pagination,
                        count=count)

@app.route("/muc7_1_5", methods=["GET","POST"]) # Danh sách xin nghỉ không lương
@login_required
def xinnghikhongluong():
    if request.method == 'GET':
        mstthuky = request.args.get("mstthuky")
        mstquanly = request.args.get("mstquanly")
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        chucvu = request.args.get("chucvu")
        chuyen = request.args.get("chuyen")
        bophan = request.args.get("bophan")
        ngay = request.args.get("ngaynghi")
        lydo = request.args.get("lydo")
        trangthai = request.args.get("trangthai")
        danhsach = laydanhsachxinnghikhongluong(mst,hoten,chucvu,chuyen,bophan,ngay,lydo,trangthai,mstquanly,mstthuky)
        count = len(danhsach)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_5.html",
                            page="Lỗi chấm công",
                            danhsach=paginated_rows,
                            pagination=pagination,
                            count=count)
    elif request.method == 'POST':
        mstquanly = request.form.get("mstquanly")
        mst = request.form.get("mst")
        hoten = request.form.get("hoten")
        chucvu = request.form.get("chucvu")
        chuyen = request.form.get("chuyen")
        bophan = request.form.get("bophan")
        ngay = request.form.get("ngaynghi")
        lydo = request.form.get("lydo")
        trangthai = request.form.get("trangthai")
        danhsach = laydanhsachxinnghikhongluong(mst,hoten,chucvu,chuyen,bophan,ngay,lydo,trangthai,mstquanly)
        data = []
        for row in danhsach:
            data.append({
                "Nhà máy": row[0],
                "Mã số thẻ": row[1],
                "Họ tên": row[2],
                "Chức danh": row[3],
                "Chuyền tổ": row[4], 
                "Phòng ban": row[5],
                "Ngày xin phép": row[6],
                "Tổng số phút": row[7],
                "Lý do": row[8],
                "Trạng thái": row[9]
            })
        df = pd.DataFrame(data)
        thoigian = datetime.now().strftime("%d%m%Y%H%M%S")
        df.to_excel(os.path.join(FOLDER_XUAT, f"xinnghikhongluong_{thoigian}.xlsx"), index=False)
        print("Tải file thành công !!!")
        return send_file(os.path.join(FOLDER_XUAT, f"xinnghikhongluong_{thoigian}.xlsx"), as_attachment=True)
        
@app.route("/muc7_1_6", methods=["GET","POST"]) # Danh sách xin nghỉ khác
@login_required
def danhsachxinnghikhac():
    if request.method == "GET":
        mstthuky = request.args.get("mstthuky")
        mstquanly = request.args.get("mstquanly")
        mst = request.args.get("mst")
        chuyen = request.args.get("chuyen")
        bophan = request.args.get("bophan")
        ngaynghi = request.args.get("ngaynghi")
        loainghi = request.args.get("loainghi")
        trangthai = request.args.get("trangthai")
        nhangiayto = request.args.get("nhangiayto")
        danhsach = laydanhsachxinnghikhac(mst,chuyen,bophan,ngaynghi,loainghi,trangthai,nhangiayto,mstthuky,mstquanly)
        count = len(danhsach)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_6.html", 
                               page="Lỗi chấm công", 
                               danhsach=paginated_rows,
                                pagination=pagination,
                                count=count)
    elif request.method == "POST":
        mstthuky = request.form.get("mstthuky")
        mstquanly = request.form.get("mstquanly")
        mst = request.form.get("mst")
        chuyen = request.form.get("chuyen")
        bophan = request.form.get("bophan")
        ngaynghi = request.form.get("ngaynghi")
        loainghi = request.form.get("loainghi")
        trangthai = request.form.get("trangthai")
        nhangiayto = request.form.get("nhangiayto")
        danhsach = laydanhsachxinnghikhac(mst,chuyen,bophan,ngaynghi,loainghi,trangthai,nhangiayto,mstthuky,mstquanly)
        data = [{
            "Nhà máy": row[0],
            "Mã số thẻ": row[1],
            "Họ tên": row[8],
            "Bộ phận": row[10],
            "Chuyền": row[9],
            "Ngày nghỉ": row[2],
            "Tổng số phút": row[3],
            "Loại nghỉ": row[4],
            "Trạng thái": row[5],
            "Nhận giấy tờ": row[6],            
        } for row in danhsach] 
        df = DataFrame(data)
        df["Mã số thẻ"] = to_numeric(df['Mã số thẻ'], errors='coerce')
        df["Ngày nghỉ"] = to_datetime(df['Ngày nghỉ'], errors='coerce')
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=xinnghikhac_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

@app.route("/muc7_1_7", methods=["GET","POST"]) # Danh sách phép tồn
@login_required
def muc7_1_7():
    if request.method == "GET":
        mst = request.args.get("mst")
        danhsach = laydanhsachphepton(mst)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_7.html", page="Lỗi chấm công",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    if request.method == "POST":
        mst = request.form.get("mst")
        danhsach = laydanhsachphepton(mst)
        result = []
        for row in danhsach:
            result.append({
                "Mã công ty": row[0],
                "Mã số thẻ": row[1],
                "Họ tên": row[2],
                "Chức danh": row[3],
                "Tháng": row[4],
                "Năm": row[5],
                "Số phút phép được dùng": row[6],
                "Số phút phép đã chốt": row[7],
                "Số phút phép chưa dùng": row[8],
                "Số phút phép cho dùng": row[9],
                "Số phút phép còn lại": row[10]
            })
        df = pd.DataFrame(result)
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=danhsachphepton_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response
           
@app.route("/muc7_1_8", methods=["GET","POST"]) # Đăng ký làm thêm giờ
@login_required
def muc7_1_8():
    
    if request.method == "GET":
        mst = request.args.get("mst")
        phongban = request.args.get("phongban")
        chuyen = request.args.get("chuyen")
        ngay = request.args.get("ngay")
        tungay = request.args.get("tungay")
        denngay = request.args.get("denngay")
        backday = True
        danhsach = laydanhsachtangca(mst,phongban,chuyen,ngay,tungay,denngay,backday)
        count = len(danhsach)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_8.html", 
                               page="Làm thêm giờ",
                               danhsach=paginated_rows,
                               pagination=pagination,
                               count=count
                               )
    elif request.method == "POST":
        data = []
        mst = request.form.get("mst")
        chuyen = request.form.get("chuyen")
        phongban = request.form.get("phongban")
        ngay = request.form.get("ngay")
        tungay = request.form.get("tungay")
        denngay = request.form.get("denngay")
        danhsach = laydanhsachtangca(mst,phongban,chuyen,ngay,tungay,denngay)
        for row in danhsach:
            data.append({
                "Nhà máy": row[0],
                "MST": row[1],
                "Họ tên": row[2],
                "Chức vụ": row[3],
                "Chuyền tổ": row[4], 
                "Phòng ban": row[5],
                "Ngày đăng ký": row[6],
                "Giờ tăng ca": row[7],
                "Giờ tăng ca thực tế": row[8]
            })
        df = pd.DataFrame(data)
        output = BytesIO()
        with ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False)

        # Điều chỉnh độ rộng cột
        output.seek(0)
        workbook = openpyxl.load_workbook(output)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        output = BytesIO()
        workbook.save(output)
        output.seek(0)
        time_stamp = datetime.now().strftime("%d%m%Y%H%M%S")
        # Trả file về cho client
        response = make_response(output.read())
        response.headers['Content-Disposition'] = f'attachment; filename=dangkytangca_{time_stamp}.xlsx'
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        return response

@app.route("/muc7_1_9", methods=["GET","POST"]) # Bảng làm thêm giờ chế độ
def muc7_1_9():
    if request.method == "GET":
        thang = int(request.args.get("thang")) if request.args.get("thang") else 0
        nam = int(request.args.get("nam")) if request.args.get("nam") else 0
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_tangcachedo(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_9.html", page="Làm thêm giờ",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        thang = request.form.get("thang") if request.form.get("thang") else datetime.now().month
        nam = request.form.get("nam") if request.form.get("nam") else datetime.now().year
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_tangcachedo(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_LAMTHEMGIO_CHEDO_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)
        sheet['A2'] = f'Tháng {thang} năm {nam}'
        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_chedo_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_chedo_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_10", methods=["GET","POST"]) # Danh sách làm thêm giờ ban ngày
def muc7_1_10():
    if request.method == "GET":
        thang = int(request.args.get("thang")) if request.args.get("thang") else 0
        nam = int(request.args.get("nam")) if request.args.get("nam") else 0
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_tangcangay(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_10.html", page="Làm thêm giờ",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        thang = request.form.get("thang") if request.form.get("thang") else datetime.now().month
        nam = request.form.get("nam") if request.form.get("nam") else datetime.now().year
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_tangcangay(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_LAMTHEMGIO_BANNGAY_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)
        sheet['A2'] = f'Tháng {thang} năm {nam}'
        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_banngay_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_banngay_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_11", methods=["GET","POST"]) # Danh sách làm thêm giờ ban đêm
def muc7_1_11():
    if request.method == "GET":
        thang = request.form.get("thang") if request.form.get("thang") else datetime.now().month
        nam = request.form.get("nam") if request.form.get("nam") else datetime.now().year
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_tangcadem(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_11.html", page="Làm thêm giờ",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        thang = request.form.get("thang")
        nam = request.form.get("nam")
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_tangcadem(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_LAMTHEMGIO_BANDEM_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)
        sheet['A2'] = f'Tháng {thang} năm {nam}'
        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_bandem_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_bandem_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_12", methods=["GET","POST"]) # Danh sách làm thêm giờ Chủ nhật
def muc7_1_12():
    if request.method == "GET":
        thang = request.form.get("thang") if request.form.get("thang") else datetime.now().month
        nam = request.form.get("nam") if request.form.get("nam") else datetime.now().year
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_tangcachunhat(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_12.html", page="Làm thêm giờ",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        thang = request.form.get("thang")
        nam = request.form.get("nam")
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_tangcachunhat(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_LAMTHEMGIO_CHUNHAT_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        sheet['A2'] = f'Tháng {thang} năm {nam}'

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_chunhat_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_chunhat_{timestamp}.xlsx"), as_attachment=True)
        
@app.route("/muc7_1_13", methods=["GET","POST"]) # Danh sách làm thêm giờ ngày lễ
def muc7_1_13():
    if request.method == "GET":
        thang = request.form.get("thang") if request.form.get("thang") else datetime.now().month
        nam = request.form.get("nam") if request.form.get("nam") else datetime.now().year
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_tangcangayle(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_13.html", page="Làm thêm giờ",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        thang = request.form.get("thang")
        nam = request.form.get("nam")
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_tangcangayle(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_LAMTHEMGIO_NGAYLE_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        sheet['A2'] = f'Tháng {thang} năm {nam}'

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_ngayle_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bang_lamthemgio_ngayle_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_14", methods=["GET","POST"]) # Bảng chấm công chi tiết chưa chốt
@login_required
def muc7_1_14():
    if request.method=="GET":
        mst = request.args.get("mst")
        chuyen = request.args.get("chuyen")
        phongban = request.args.get("phongban")
        tungay = request.args.get("tungay")
        denngay = request.args.get("denngay")
        phanloai = request.args.get("phanloai")
        rows = laydanhsachchamcong(mst,chuyen,phongban,tungay,denngay,phanloai)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_14.html", page="Bảng chấm công",
                            danhsach=paginated_rows, 
                            pagination=pagination,
                            count=count)
    elif request.method=="POST":
        mst = request.form.get('mst')
        chuyen = request.form.get('chuyen')
        phongban = request.form.get('phongban')
        tungay = request.form.get("tungay")
        denngay = request.form.get("denngay")
        phanloai = request.form.get("phanloai")
        danhsach = laydanhsachchamcong(mst,chuyen,phongban,tungay,denngay,phanloai)
        workbook = openpyxl.load_workbook(FILE_MAU_BANGCONG_CHUACHOT_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'
        
        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-1]]
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # Duyệt qua các ô trong khu vực G4:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này            

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chuachot_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chuachot_{timestamp}.xlsx"), as_attachment=True)
                
@app.route("/muc7_1_15", methods=["GET","POST"]) # Bảng chấm công chi tiết chốt
@login_required
def muc7_1_15():
    if request.method=="GET":
        mst = request.args.get("mst")
        chuyen = request.args.get('chuyen')
        phongban = request.args.get("phongban")
        tungay = request.args.get("tungay")
        denngay = request.args.get("denngay")
        phanloai = request.args.get("phanloai")
        rows = laydanhsachchamcongchot(mst,chuyen,phongban,tungay,denngay,phanloai)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        danhsachphongban = laycacphongban()
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_15.html", page="Bảng chấm công",
                            danhsach=paginated_rows, 
                            pagination=pagination,
                            count=count,
                            danhsachphongban=danhsachphongban)
    elif request.method=="POST":
        mst = request.form.get('mst')
        chuyen = request.form.get('chuyen')
        phongban = request.form.get('phongban')
        tungay = request.form.get("tungay")
        denngay = request.form.get("denngay")
        phanloai = request.form.get("phanloai")
        danhsach = laydanhsachchamcongchot(mst,chuyen,phongban,tungay,denngay,phanloai)
        workbook = openpyxl.load_workbook(FILE_MAU_BANGCONG_CHOT_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 50000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-1]]
            data[7] = datetime.strptime(data[7],"%Y-%m-%d")
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        number_style = NamedStyle(name="number_style", number_format="0")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 50001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chot_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chot_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_16", methods=["GET","POST"]) # Bảng chấm công hành chính
@login_required
def muc7_1_16():
    
    if request.method == "GET":
        thang = int(request.args.get("thang")) if request.args.get("thang") else 0
        nam = int(request.args.get("nam")) if request.args.get("nam") else 0
        mst = request.args.get("mst")
        bophan = request.args.get("bophan")
        chuyen = request.args.get("chuyen")
        danhsach = lay_bangcong_kx(thang,nam,mst,bophan,chuyen)
        total = len(danhsach)
        page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 15
        start = (page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_16.html", page="Bảng chấm công",
                                danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
        
    elif request.method == "POST":
        thang = request.form.get("thang")
        nam = request.form.get("nam")
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_bangcong_kx(thang,nam,mst,bophan,chuyen)
        workbook = openpyxl.load_workbook(FILE_MAU_BANGCONG_HANHCHINH_KX)

        sheet = workbook['BẢNG CHẤM CÔNG HÀNH CHÍNH']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        sheet['A2'] = f'Tháng {thang} năm {nam}'

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 10000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này          

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_hanhchinh_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_hanhchinh_{timestamp}.xlsx"), as_attachment=True)
    
@app.route("/muc7_1_17", methods=["GET","POST"]) # Bảng chấm công tổng hợp
def muc7_1_17():
    if request.method == "GET":
        try:
            thang = int(request.args.get("thang")) if request.args.get("thang") else 0
            nam = int(request.args.get("nam")) if request.args.get("nam") else 0
            mst = request.args.get("mst")
            bophan = request.args.get("bophan")
            chuyen = request.args.get("chuyen")
            danhsach = lay_bangcongthang_kx(mst,bophan,chuyen,thang,nam)
            total = len(danhsach)
            page = request.args.get(get_page_parameter(), type=int, default=1)
            per_page = 15
            start = (page - 1) * per_page
            end = start + per_page
            paginated_rows = danhsach[start:end]
            pagination = Pagination(page=page, per_page=per_page, total=total, css_framework='bootstrap4')
            return render_template("7_1_17.html", page="Bảng chấm công",
                                    danhsach=paginated_rows, 
                                    pagination=pagination,
                                    count=total)
        except Exception as e:
            print(e)
            return render_template("7_1_17.html",
                                    danhsach=[])
    else:
        thang = int(request.form.get("thang")) if request.args.get("thang") else 0
        nam = int(request.form.get("nam")) if request.args.get("nam") else 0
        mst = request.form.get("mst")
        bophan = request.form.get("bophan")
        chuyen = request.form.get("chuyen")
        danhsach = lay_bangcongthang_kx(mst,bophan,chuyen,thang,nam)
        workbook = openpyxl.load_workbook(FILE_MAU_BANGCONG_TONGHOP)

        sheet = workbook['BẢNG CHẤM CÔNG TỔNG HỢP']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        sheet['A2'] = f'Tháng {thang} năm {nam}'

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(6, 10000 - 6 + 1)

        for row in danhsach:
            data = [y for y in row[:-3]]
            data[6] = datetime.strptime(data[6],"%Y-%m-%d") if data[6] else ""
            data[7] = datetime.strptime(data[7],"%Y-%m-%d") if data[7] else ""
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        number_style = NamedStyle(name="number_style", number_format="0.00")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(6, 10001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['G', 'H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này
            for col in ['J', 'K','L', 'M','N', 'O','P', 'Q','R', 'S','T', 'U', 'X','Y', 'Z','AA','AB', 'AC','AD', 'AE', 'AF','AG', 'AH','AI', 'AJ', 'AK','AL', 'AM', 'AN']:
                cell = sheet[f"{col}{row}"]
                if cell.value and int(cell.value) > 0:
                    try:
                        cell.style = number_style
                    except ValueError:
                        pass  # Nếu giá trị không phải là ngày, bỏ qua ô này
            

        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_tonghop_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_tonghop_{timestamp}.xlsx"), as_attachment=True)

@app.route("/muc7_1_18", methods=["GET","POST"]) # Bảng chấm công chi tiết chốt quá khứ
@login_required
def muc7_1_18():
    if request.method=="GET":
        mst = request.args.get("mst")
        chuyen = request.args.get('chuyen')
        phongban = request.args.get("phongban")
        tungay = request.args.get("tungay")
        denngay = request.args.get("denngay")
        phanloai = request.args.get("phanloai")
        rows = laydanhsachchamcongchotquakhu(mst,chuyen,phongban,tungay,denngay,phanloai)
        count = len(rows)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(rows)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = rows[start:end]
        danhsachphongban = laycacphongban()
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("7_1_18.html", page="Bảng chấm công",
                            danhsach=paginated_rows, 
                            pagination=pagination,
                            count=count,
                            danhsachphongban=danhsachphongban)
    elif request.method=="POST":
        mst = request.form.get('mst')
        chuyen = request.form.get('chuyen')
        phongban = request.form.get('phongban')
        tungay = request.form.get("tungay")
        denngay = request.form.get("denngay")
        phanloai = request.form.get("phanloai")
        danhsach = laydanhsachchamcongchotquakhu(mst,chuyen,phongban,tungay,denngay,phanloai)
        workbook = openpyxl.load_workbook(FILE_MAU_BANGCONG_CHOT_KX)

        sheet = workbook['Sheet1']  # Thay 'Sheet1' bằng tên sheet của bạn
        image_path = HINHANH_LOGO
        # Tạo đối tượng hình ảnh
        img = Image(image_path)
        # Điều chỉnh kích thước hình ảnh xuống 70% so với kích thước gốc
        img.width = img.width * 0.25
        img.height = img.height * 0.25

        # Di chuyển ảnh: anchor vào ô A2 và điều chỉnh tọa độ di chuyển
        img.anchor = 'A1'

        # Chèn hình ảnh vào sheet
        sheet.add_image(img)

        # Xóa hàng từ hàng 7 đến hàng 10000
        sheet.delete_rows(4, 50000 - 4 + 1)

        for row in danhsach:
            data = [y for y in row[:-1]]
            data[7] = datetime.strptime(data[7],"%Y-%m-%d")
            sheet.append(data)

        # Tạo kiểu định dạng ngày
        date_style = NamedStyle(name="date_style", number_format="DD/MM/YYYY")
        # Duyệt qua các ô trong khu vực G7:H10000
        for row in range(4, 50001):  # Bắt đầu từ dòng 7 đến dòng 10000
            for col in ['H']:
                cell = sheet[f"{col}{row}"]
                
                try:
                    cell.style = date_style
                except ValueError:
                    pass  # Nếu giá trị không phải là ngày, bỏ qua ô này
            
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        workbook.save(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chot_{timestamp}.xlsx"))
        return send_file(os.path.join(os.path.dirname(__file__),f"nhapxuat/xuat/bangchamcong_chitiet_chot_{timestamp}.xlsx"), as_attachment=True)
            
@app.route("/muc8_1", methods=["GET","POST"])
@login_required
def ykienkhieunai():

    return render_template("8_1.html", page="8.1 Danh sách ý kiến khiếu nại")

@app.route("/muc8_2", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def capnhatykienkhieunai():

    return render_template("8_2.html", page="8.2 Cập nhật ý kiến khiếu nại")
    
@app.route("/muc9_1", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def xulykiluat():
    
    if request.method == "GET":
        danhsach = laydanhsachkyluat()
        return render_template("9_1.html", page="9.1 Xử lý kỉ luật",danhsach=danhsach)
    else:
        try:
            mst = request.form.get("mst")
            if not mst:
                flash("Chưa có thông tin người vi phạm")
                return redirect("/muc9_1") 
            hoten = request.form.get("hoten")
            chucvu = request.form.get("chucvu")
            bophan = request.form.get("bophan")
            chuyento = request.form.get("chuyento")
            ngayvao = request.form.get("ngayvao")
            ngayvipham = request.form.get("ngayvipham")
            diadiem = request.form.get("diadiem")
            ngaylapbienban = request.form.get("ngaylapbienban")
            noidung = request.form.get("noidung")
            bienphap = request.form.get("bienphap")
            cacanhvipham = request.files.getlist("file_anh")
            bienbankiluat = request.files.get("file_bienban") 
            os.makedirs(os.path.join(FOLDER_BIENBAN,f"{mst}_{ngayvipham}"),exist_ok=True)
            
            for anh in cacanhvipham:
                anh.save(os.path.join(FOLDER_BIENBAN,f"{mst}_{ngayvipham}",f"{cacanhvipham.index(anh,start=1)}.jpg"))  
            bienbankiluat.save(os.path.join(FOLDER_BIENBAN,f"{mst}_{ngayvipham}"),"bienban.pdf")
            if themdanhsachkyluat(mst,hoten,chucvu,bophan,chuyento,ngayvao,ngayvipham,diadiem,ngaylapbienban,noidung,bienphap):
                flash("Thêm biên bản kỷ luật thành công !!!")
            else:
                flash("Thêm biên bản kỷ luật thất bại !!!")
        except Exception as e:
            print(f"Thêm biên bản kỷ luật thất bại {e}!!!")
        return redirect("/muc9_1") 
    
@app.route("/muc10_1", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def phongvannghiviec():
        
    return render_template("10_1.html", page="10.1 Tổng hợp phỏng vấn nghỉ việc")

@app.route("/muc10_2", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def nhandonnghiviec():
    if request.method == "GET":
        mst = request.args.get("mst")
        hoten = request.args.get("hoten")
        chuyen = request.args.get("chuyen")
        phongban = request.args.get("phongban")
        ngaynopdon = request.args.get("ngaynopdon")
        ngaynghi = request.args.get("ngaynghi")
        sapdenhan = request.args.get("sapdenhan")
        danhsach = laydanhsach_chonghiviec(mst,hoten,chuyen,phongban,ngaynopdon,ngaynghi,sapdenhan)
        current_page = request.args.get(get_page_parameter(), type=int, default=1)
        per_page = 10
        total = len(danhsach)
        start = (current_page - 1) * per_page
        end = start + per_page
        paginated_rows = danhsach[start:end]
        pagination = Pagination(page=current_page, per_page=per_page, total=total, css_framework='bootstrap4')
        return render_template("10_2.html", 
                            page="10.2 Tổng hợp đơn nghỉ việc",
                            danhsach=paginated_rows, 
                                pagination=pagination,
                                count=total)
    elif request.method == "POST":
        mst = request.form.get("form_manhanvien")
        hoten = request.form.get("form_hovaten")
        chucdanh = request.form.get("form_chucvu")
        chuyen = request.form.get("form_chuyento")
        phongban = request.form.get("form_bophan")
        ngaynopdon = request.form.get("form_ngaynopdon")
        ngaynghi = request.form.get("form_ngaydukiennghi")
        ghichu = request.form.get("form_ghichu")
        if themdonxinnghi(mst,hoten,chucdanh,chuyen,phongban,ngaynopdon,ngaynghi,ghichu):
            print("Thêm đơn xin nghỉ thành công !!!")
        else:
            print("Thêm đơn xin nghỉ thất bại !!!")
        return redirect(f"/muc10_2?mst={mst}")
    
@app.route("/muc10_3", methods=["GET","POST"])
@login_required
@roles_required('hr','sa','gd')
def inchamduthopdong():
     
    if request.method == "GET":
        return render_template("10_3.html", page="10.3 In chấm dứt hợp đồng")
    elif request.method == "POST":
        mst = request.form.get("mst")
        ngaylamhopdong = datetime.strptime(request.form.get("ngaylamhd"),"%Y-%m-%d").strftime("%d")
        thanglamhopdong = datetime.strptime(request.form.get("ngaylamhd"),"%Y-%m-%d").strftime("%m")
        namlamhopdong = datetime.strptime(request.form.get("ngaylamhd"),"%Y-%m-%d").strftime("%Y")
        tennhanvien = request.form.get("hoten")
        chucvu = request.form.get("chucvu")
        ngaynghi = datetime.strptime(request.form.get("ngaynghi"),"%Y-%m-%d").strftime("%d/%m/%Y")
        ngaysinh = request.form.get("ngaysinh")
        diachi = request.form.get("diachi")
        bophan = request.form.get("bophan")
        lydo = request.form.get("lydo")
        try:
            file = inchamduthd(mst,
                ngaylamhopdong,
                thanglamhopdong,
                namlamhopdong,
                tennhanvien,
                chucvu,
                ngaynghi,
                ngaysinh,
                diachi,
                bophan,
                lydo)
            if file:
                return send_file(file, as_attachment=True, download_name="chamduthopdong.xlsx")
            else:

                return redirect("/muc10_3")
        except Exception as e:
            print(e)
            return redirect("/muc10_3") 
        
@app.route("/muc12", methods=["GET","POST"])
@login_required
def khong_kiem_xuong():
    try:
        if request.method=="GET":
            return render_template("12.html")
        else:
            return "OK"
    except Exception as e:
        print(e)
        return "NOT OK"

@app.route("/admin", methods=["GET"])
@login_required
@roles_required('sa')
def admin_page():
    trangthai = trang_thai_function_12()
    return render_template("admin.html",trangthai=trangthai)